#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""M3/M4 测试：重算降级与错误扫描、中文样式、数字格式、图表。"""

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook

from duduexcel import recalc, styling

TMP = Path(__file__).parent / "_tmp_m34.xlsx"


def check(name, cond, extra="") -> bool:
    print(f"[{'PASS' if cond else 'FAIL'}] {name} {extra}")
    return cond


def build() -> Path:
    if TMP.exists():
        TMP.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["月份", "销售额"])
    for m, v in [("1月", 120), ("2月", 180), ("3月", 90), ("4月", 210)]:
        ws.append([m, v])
    # 故意造一个除零错误用于错误扫描验证
    ws["D1"] = "=1/0"
    wb.save(TMP)
    return TMP


def main() -> int:
    p = build()
    ok = True

    # --- M3：环境缺失时的优雅降级（不静默成功）---
    r = recalc.recalculate(p)
    soffice_missing = r.get("reason") == "missing_libreoffice"
    ok &= check("无 LibreOffice 时明确降级", soffice_missing and r["ok"] is False,
                f"-> reason={r.get('reason')}")
    if soffice_missing:
        ok &= check("降级时给出可操作提示", "LibreOffice" in r.get("message", ""))

    # 外链熔断函数在无外链时返回空（不误报）
    at_risk = recalc.external_links_at_risk(p)
    ok &= check("无外链时不误报风险", at_risk == [], f"-> {at_risk}")

    # --- M3：错误扫描（读回的缓存值里含 #DIV/0!）---
    # 注：openpyxl 刚写入未重算时公式无缓存值，故先造一个带错误值的文件
    p2 = Path(str(TMP).replace("_tmp_m34", "_tmp_err"))
    if p2.exists():
        p2.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "#DIV/0!"
    ws["A2"] = "#NAME?"
    ws["A3"] = "正常值"
    wb.save(p2)
    scan = recalc.scan_formula_errors(p2)
    ok &= check("错误扫描识别 2 个错误", scan["total_errors"] == 2,
                f"-> total={scan['total_errors']}, types={list(scan['errors_by_type'].keys())}")
    ok &= check("错误按类型分组且含位置",
                "S!A1" in scan["errors_by_type"].get("#DIV/0!", {}).get("locations", []),
                f"-> {scan['errors_by_type'].get('#DIV/0!')}")

    # --- M4：中文样式 ---
    res = styling.apply_chinese_style(p, "数据")
    ok &= check("中文样式套用成功", res["styled_rows"] >= 5 and res["styled_columns"] >= 2,
                f"-> {res['styled_rows']}行 x {res['styled_columns']}列, font={res['font']}")

    wb = load_workbook(p)
    ws = wb["数据"]
    ok &= check("表头字体为中文字体", ws["A1"].font.name in styling.CJK_FONTS,
                f"-> {ws['A1'].font.name}")
    ok &= check("表头已冻结", ws.freeze_panes is not None, f"-> {ws.freeze_panes}")
    ok &= check("列宽已自适应（不低于最小宽度 8）",
                (ws.column_dimensions["A"].width or 0) >= 8,
                f"-> A列宽={ws.column_dimensions['A'].width}")
    # 用长文本列验证列宽确实随内容增长（中文按 2 字符宽）
    ws["E1"] = "这是一列很长的中文表头标题"
    wb.save(p)
    styling.apply_chinese_style(p, "数据", header_row=1)
    wb2 = load_workbook(p)
    w_long = wb2["数据"].column_dimensions["E"].width or 0
    ok &= check("长中文表头列宽显著变宽", w_long > 12, f"-> E列宽={w_long}")
    wb2.close()
    wb.close()

    # --- M4：数字格式 ---
    nf = styling.set_number_format(p, "数据", "B2:B5", "cny")
    ok &= check("人民币格式套用", "¥" in nf["number_format"], f"-> {nf['number_format']}")
    ok &= check("格式作用到 4 个单元格", nf["cells_updated"] == 4, f"-> {nf['cells_updated']}")

    # 内置名与自定义码都可
    nf2 = styling.set_number_format(p, "数据", "B2:B3", "0.00%")
    ok &= check("自定义格式码可用", nf2["number_format"] == "0.00%")

    # --- M4：图表 ---
    ch = styling.add_chart(p, "数据", chart_type="bar", data_range="B2:B5",
                           categories_range="A2:A5", title="月度销售额", anchor_cell="E2")
    ok &= check("柱状图写入成功", ch["chart_type"] == "bar")
    wb = load_workbook(p)
    ok &= check("工作簿中确实存在图表", len(wb["数据"]._charts) >= 1,
                f"-> charts={len(wb['数据']._charts)}")
    wb.close()

    # 折线图与饼图
    ok &= check("折线图写入成功",
                styling.add_chart(p, "数据", chart_type="line", data_range="B2:B5",
                                  anchor_cell="E20")["chart_type"] == "line")
    ok &= check("饼图写入成功",
                styling.add_chart(p, "数据", chart_type="pie", data_range="B2:B5",
                                  categories_range="A2:A5", anchor_cell="E38")["chart_type"] == "pie")

    # 非法图表类型应报错
    try:
        styling.add_chart(p, "数据", chart_type="雷达图", data_range="B2:B5")
        ok &= check("非法图表类型应报错", False)
    except ValueError as e:
        ok &= check("非法图表类型给出可用清单", "bar" in str(e), f"-> {str(e)[:50]}")

    # 清理
    for f in (TMP, p2):
        if f.exists():
            f.unlink()
        bak = Path(str(f) + ".bak")
        if bak.exists():
            bak.unlink()

    print("")
    print(f"===== M3/M4 测试: {'全部通过' if ok else '存在失败'} =====")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
