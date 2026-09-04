#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M8：公式重算真实闭环测试（需要本机安装 LibreOffice）。

验证链路：写公式（openpyxl 无缓存值）→ recalculate → 读回真实计算结果。
同时验证安全护栏：外链熔断、公式错误扫描。

若未安装 LibreOffice，本测试会**跳过**真实重算部分（不失败），
但仍验证降级行为正确（明确告知而非静默假装成功）。
"""

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook

from duduexcel import recalc

TMP = Path(__file__).parent / "_tmp_m8.xlsx"
ERR = Path(__file__).parent / "_tmp_m8_err.xlsx"


def check(name, cond, extra="") -> bool:
    print(f"[{'PASS' if cond else 'FAIL'}] {name} {extra}")
    return cond


def build() -> Path:
    if TMP.exists():
        TMP.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "计算"
    ws.append(["单价", "数量", "金额"])
    ws.append([100, 3, "=A2*B2"])
    ws.append([50, 4, "=A3*B3"])
    ws.append([None, None, "=SUM(C2:C3)"])
    wb.save(TMP)
    return TMP


def main() -> int:
    ok = True
    soffice = recalc._find_soffice()

    if soffice is None:
        print("⏭  未检测到 LibreOffice，跳过真实重算测试\n")
        # 仍验证降级行为
        p = build()
        r = recalc.recalculate(p)
        ok &= check("无 LibreOffice 时明确降级", r.get("reason") == "missing_libreoffice",
                    f"-> {r.get('reason')}")
        ok &= check("降级给出安装指引", "LibreOffice" in r.get("message", ""))
        if TMP.exists():
            TMP.unlink()
        print(f"\n===== M8 测试（降级路径）: {'全部通过' if ok else '存在失败'} =====")
        return 0 if ok else 1

    print(f"LibreOffice: {soffice}\n")
    p = build()

    # 1. 重算前：openpyxl 写的公式没有缓存值
    wb = load_workbook(p, data_only=True)
    ok &= check("重算前公式无缓存值", wb["计算"]["C2"].value is None,
                f"-> C2={wb['计算']['C2'].value}")
    wb.close()

    # 2. 执行重算
    r = recalc.recalculate(p, timeout=180)
    ok &= check("重算成功", r.get("recalculated") is True,
                f"-> ok={r.get('ok')}, 耗时={r.get('elapsed_seconds')}s")
    ok &= check("文件被重写", r.get("file_rewritten") is True)
    ok &= check("无公式错误", r.get("total_errors") == 0,
                f"-> total_errors={r.get('total_errors')}")

    # 3. 重算后：读到真实结果
    if r.get("recalculated"):
        wb = load_workbook(p, data_only=True)
        c2 = wb["计算"]["C2"].value
        c3 = wb["计算"]["C3"].value
        c4 = wb["计算"]["C4"].value
        wb.close()
        ok &= check("C2 = 单价×数量 = 300", c2 == 300, f"-> {c2}")
        ok &= check("C3 = 单价×数量 = 200", c3 == 200, f"-> {c3}")
        ok &= check("C4 = SUM 汇总 = 500", c4 == 500, f"-> {c4}")

    # 4. 公式错误扫描（造一个含错误的文件）
    if ERR.exists():
        ERR.unlink()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=1/0"      # 会被识别为 #DIV/0!（若已重算）；未重算则读回 None
    ws["A2"] = "正常"
    wb.save(ERR)
    # 先重算让错误值落盘
    recalc.recalculate(ERR, timeout=180)
    scan = recalc.scan_formula_errors(ERR)
    ok &= check("错误扫描能识别 #DIV/0!", 
                scan.get("total_errors", 0) >= 1,
                f"-> total={scan.get('total_errors')}, types={list(scan.get('errors_by_type', {}).keys())}")

    # 5. 无外链文件不应误报风险
    at_risk = recalc.external_links_at_risk(p)
    ok &= check("无外链时不误报", at_risk == [], f"-> {at_risk}")

    for f in (TMP, ERR):
        if f.exists():
            f.unlink()
        b = Path(str(f) + ".bak")
        if b.exists():
            b.unlink()

    print("")
    print(f"===== M8 测试: {'全部通过' if ok else '存在失败'} =====")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
