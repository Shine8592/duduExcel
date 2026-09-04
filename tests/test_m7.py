#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M7 测试：汲取 excel-vision-mcp 后新增的能力。

1. 格式语义标记：[B]/[I]/[S]/[HL:色]/[C:色]/[M]，且朴素表零开销
2. 隐藏行列：默认跳过 + 报告数量 + 被公式引用的保留（HIDDEN-REF）
3. 内嵌图片清单（零依赖 zipfile 扫描）
4. 原子保存：写入失败时原文件不受影响
5. 多目录沙箱（; 分隔）
"""

import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from duduexcel import excel_ops
from duduexcel.safety import SafetyError, resolve_path

TMP = Path(__file__).parent / "_tmp_m7.xlsx"


def check(name, cond, extra="") -> bool:
    print(f"[{'PASS' if cond else 'FAIL'}] {name} {extra}")
    return cond


def build_format_sheet() -> Path:
    """造一个带格式语义的表：删除线=已取消、黄底=待审阅、粗体表头。"""
    if TMP.exists():
        TMP.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "规格"
    ws.append(["特性", "状态"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.append(["车辆调度", "已批准"])
    ws.append(["旧版导出", "已取消"])
    ws["A3"].font = Font(strike=True)          # 删除线 → 已取消
    ws["B3"].font = Font(strike=True)
    ws.append(["司机点名", "待审阅"])
    ws["B4"].fill = PatternFill("solid", fgColor="FFFF00")  # 黄底 → 待审阅
    wb.save(TMP)
    return TMP


def main() -> int:
    ok = True
    p = build_format_sheet()

    # --- 1. 格式语义标记 ---
    r = excel_ops.read_range(p, sheet="规格", limit=10)
    fm = r.get("format_markers", {})
    ok &= check("检测到格式标记", bool(fm), f"-> {list(fm.items())[:4]}")
    ok &= check("粗体表头被标记 [B]", fm.get("A1") == ["B"], f"-> A1={fm.get('A1')}")
    ok &= check("删除线被标记 [S]（语义：已取消）",
                "S" in (fm.get("A3") or []), f"-> A3={fm.get('A3')}")
    ok &= check("黄底被标记 [HL:yellow]（语义：待审阅）",
                any("HL:" in m for m in (fm.get("B4") or [])), f"-> B4={fm.get('B4')}")
    ok &= check("提供标记图例", "marker_legend" in r)
    ok &= check("提供格式化视图", "formatted_view" in r)
    print("      格式化视图示例:")
    for line in r.get("formatted_view", [])[:4]:
        print(f"        {line}")

    # --- 2. 朴素表零开销（无格式则不产生标记）---
    plain = Path(__file__).parent / "_tmp_plain.xlsx"
    if plain.exists():
        plain.unlink()
    wb = Workbook()
    ws = wb.active
    ws.append(["列1", "列2"])
    ws.append(["a", 1])
    wb.save(plain)
    rp = excel_ops.read_range(plain, limit=10)
    ok &= check("朴素表不产生格式标记（零额外 token）",
                "format_markers" not in rp, f"-> keys含format_markers={'format_markers' in rp}")
    if plain.exists():
        plain.unlink()

    # --- 3. 隐藏行列处理 ---
    p2 = Path(__file__).parent / "_tmp_hidden.xlsx"
    if p2.exists():
        p2.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["项目", "数量", "单价"])
    for row in [("苹果", 10, 5), ("隐藏项", 99, 99), ("香蕉", 20, 3)]:
        ws.append(list(row))
    ws["D1"] = "合计"
    ws["D2"] = "=B2*C2+B4*C4"   # 可见公式，引用了可见行
    ws.row_dimensions[3].hidden = True   # 隐藏"隐藏项"行
    wb.save(p2)

    rh = excel_ops.read_range(p2, sheet="数据", limit=10)   # 默认跳过隐藏
    hs = rh.get("hidden_skipped")
    ok &= check("隐藏行被跳过并报告", hs is not None and hs["rows"] >= 1,
                f"-> {hs}")
    ok &= check("隐藏时给出不静默丢弃的说明", "note" in (hs or {}),
                f"-> {str((hs or {}).get('note'))[:60]}")

    rh2 = excel_ops.read_range(p2, sheet="数据", limit=10, include_hidden=True)
    ok &= check("include_hidden=true 时读到隐藏行",
                len(rh2["grid"]) > len(rh["grid"]),
                f"-> 隐藏时{len(rh['grid'])}行 vs 全读{len(rh2['grid'])}行")
    if p2.exists():
        p2.unlink()

    # --- 4. 内嵌图片清单 ---
    p3 = Path(__file__).parent / "_tmp_img.xlsx"
    if p3.exists():
        p3.unlink()
    wb = Workbook()
    wb.active["A1"] = "带图的表"
    wb.save(p3)
    # 手工往 xlsx（zip）里塞一张 PNG，模拟内嵌图片
    import zipfile
    png = (b"\x89PNG\r\n\x1a\n" + b"\x00" * 8
           + (100).to_bytes(4, "big") + (50).to_bytes(4, "big") + b"\x00" * 20)
    with zipfile.ZipFile(p3, "a") as z:
        z.writestr("xl/media/image1.png", png)
    li = excel_ops.list_images(p3)
    ok &= check("扫描到内嵌图片", li.get("count") == 1, f"-> count={li.get('count')}")
    if li.get("images"):
        img = li["images"][0]
        ok &= check("解析出图片尺寸", img.get("width") == 100 and img.get("height") == 50,
                    f"-> {img.get('width')}x{img.get('height')}")
    # workbook_info 也应带图片数
    info = excel_ops.inspect_workbook(p3)
    ok &= check("workbook_info 报告图片数", info.get("embedded_images") == 1,
                f"-> {info.get('embedded_images')}")
    if p3.exists():
        p3.unlink()

    # --- 5. 原子保存：写入失败时原文件完好 ---
    p4 = Path(__file__).parent / "_tmp_atomic.xlsx"
    if p4.exists():
        p4.unlink()
    wb = Workbook()
    wb.active["A1"] = "原始内容"
    wb.save(p4)
    before = p4.read_bytes()
    try:
        # 故意写一个会失败的调用（不存在的 sheet 且不允许新建 → 抛错，不应写坏文件）
        excel_ops.write_cells(p4, sheet="不存在", cells=[{"cell": "A1", "value": "x"}])
    except Exception:
        pass
    after = p4.read_bytes()
    ok &= check("写入失败后原文件未被破坏", before == after,
                f"-> 文件大小 {len(before)} vs {len(after)}")
    if p4.exists():
        p4.unlink()

    # --- 6. 多目录沙箱 ---
    d1 = Path(__file__).parent
    d2 = Path(__file__).parent.parent
    os.environ["DUDU_EXCEL_ROOT"] = f"{d1};{d2}"
    try:
        ok &= check("多目录沙箱：第一个目录内的文件可访问",
                    resolve_path("test_m7.py").exists())
        ok &= check("多目录沙箱：第二个目录内的文件也可访问",
                    resolve_path("README.md").exists())
        try:
            resolve_path("../../Windows/System32/config/SAM")
            ok &= check("多目录沙箱：越界仍被拒绝", False)
        except SafetyError:
            ok &= check("多目录沙箱：越界仍被拒绝", True)
    finally:
        os.environ.pop("DUDU_EXCEL_ROOT", None)

    if TMP.exists():
        TMP.unlink()
    for f in Path(__file__).parent.glob("_tmp_*.bak"):
        f.unlink()

    print("")
    print(f"===== M7 测试: {'全部通过' if ok else '存在失败'} =====")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
