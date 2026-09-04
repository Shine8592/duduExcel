#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""M6 测试：透视表、条件格式、多表关联。"""

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook

from duduexcel import advanced

TMP = Path(__file__).parent / "_tmp_m6.xlsx"


def check(name, cond, extra="") -> bool:
    print(f"[{'PASS' if cond else 'FAIL'}] {name} {extra}")
    return cond


def build() -> Path:
    if TMP.exists():
        TMP.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "销售"
    ws.append(["产品", "部门", "销售额"])
    for r in [("笔记本", "研发", 120), ("鼠标", "研发", 35), ("显示器", "市场", 89),
              ("键盘", "研发", 45), ("耳机", "销售", 210), ("音箱", "销售", 76)]:
        ws.append(list(r))

    # v2 表：改了鼠标销售额、新增投影仪、删掉音箱
    ws2 = wb.create_sheet("销售v2")
    ws2.append(["产品", "部门", "销售额"])
    for r in [("笔记本", "研发", 120), ("鼠标", "研发", 99), ("显示器", "市场", 89),
              ("键盘", "研发", 45), ("耳机", "销售", 210), ("投影仪", "市场", 320)]:
        ws2.append(list(r))

    # 部门维度表用于 join
    ws3 = wb.create_sheet("部门表")
    ws3.append(["部门", "负责人"])
    for r in [("研发", "张三"), ("市场", "李四"), ("销售", "王五")]:
        ws3.append(list(r))
    wb.save(TMP)
    return TMP


def main() -> int:
    p = build()
    ok = True

    # --- 透视表 ---
    pv = advanced.create_pivot(p, "销售", rows=["部门"], values=["销售额"], agg_func="sum")
    ok &= check("透视表生成", pv["result_rows"] == 3, f"-> {pv['result_rows']} 行分组")
    ok &= check("透视表诚实标注为非交互", pv["is_interactive_pivot"] is False)

    wb = load_workbook(p)
    ok &= check("透视表写入新工作表", "透视表" in wb.sheetnames, f"-> {wb.sheetnames}")
    ws = wb["透视表"]
    # 研发 = 120+35+45 = 200
    vals = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)}
    ok &= check("透视表聚合值正确", vals.get("研发") == 200, f"-> {vals}")
    wb.close()

    # 交叉表（带 columns）
    pv2 = advanced.create_pivot(p, "销售", rows=["部门"], values=["销售额"],
                                columns=["产品"], target_sheet="交叉表")
    ok &= check("交叉表生成", pv2["result_rows"] == 3 and pv2["columns"] == ["产品"],
                f"-> rows={pv2['result_rows']}")

    # --- 条件格式 ---
    cf = advanced.add_conditional_format(p, "销售", "C2:C7", "greater_than", value=100)
    ok &= check("条件格式 greater_than 生效", cf["cond_type"] == "greater_than")
    wb = load_workbook(p)
    ok &= check("条件格式规则已写入",
                len(wb["销售"].conditional_formatting._cf_rules) >= 1,
                f"-> 规则区数={len(wb['销售'].conditional_formatting._cf_rules)}")
    wb.close()

    ok &= check("数据条类型可用",
                advanced.add_conditional_format(p, "销售", "C2:C7", "data_bar")["cond_type"] == "data_bar")
    ok &= check("色阶类型可用",
                advanced.add_conditional_format(p, "销售", "C2:C7", "color_scale")["cond_type"] == "color_scale")
    ok &= check("between 类型可用",
                advanced.add_conditional_format(p, "销售", "C2:C7", "between",
                                                value=50, value2=150)["cond_type"] == "between")
    ok &= check("重复值高亮可用",
                advanced.add_conditional_format(p, "销售", "A2:A7", "duplicate")["cond_type"] == "duplicate")

    # 错误类型
    try:
        advanced.add_conditional_format(p, "销售", "C2:C7", "不存在类型")
        ok &= check("错误条件类型应报错", False)
    except Exception as e:
        ok &= check("错误条件类型给出可用清单", "可用" in str(e), f"-> {str(e)[:50]}")

    # --- 多表比较 ---
    cmp = advanced.compare_sheets(p, "销售", "销售v2", key_column="产品")
    ok &= check("比较：仅在左表 1 条（音箱）", cmp["only_in_left"] == 1,
                f"-> only_left={cmp['only_in_left_sample']}")
    ok &= check("比较：仅在右表 1 条（投影仪）", cmp["only_in_right"] == 1,
                f"-> only_right={cmp['only_in_right_sample']}")
    ok &= check("比较：值差异 1 处（鼠标 35→99）", cmp["value_differences"] == 1,
                f"-> {cmp['differences']}")

    # --- 多表关联 ---
    jn = advanced.join_sheets(p, "销售", "部门表", on="部门", how="left", limit=3)
    ok &= check("左连接返回结果", jn["returned"] == 3, f"-> {jn['returned']} 行")
    ok &= check("连接结果含右表列", "负责人" in jn["columns"], f"-> {jn['columns']}")
    ok &= check("连接自报 token 节省", "_meta" in jn)

    # 错误：关联键不存在
    try:
        advanced.join_sheets(p, "销售", "部门表", on="不存在的列")
        ok &= check("错误关联键应报错", False)
    except Exception as e:
        ok &= check("错误关联键给出两表列名", "必须在两个表中都存在" in str(e), f"-> {str(e)[:60]}")

    if TMP.exists():
        TMP.unlink()
    bak = Path(str(TMP) + ".bak")
    if bak.exists():
        bak.unlink()

    print("")
    print(f"===== M6 测试: {'全部通过' if ok else '存在失败'} =====")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
