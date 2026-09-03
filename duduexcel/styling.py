#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M4：中文场景样式预设 + 图表（差异化能力）。

## 为什么这是差异化
调研发现两个空白：
1. **官方 xlsx skill 正文零图表指导**（"charting" 只出现在 description 触发词里）
2. **竞品 knorq 明确不支持图表/透视表/条件格式**（写在 Known Limitations 里）
3. **中文场景无人覆盖**：官方规范要求"专业字体（Arial / Times New Roman）"，
   但这对中文表格并不适用 —— 中文需要微软雅黑/宋体，
   货币格式需要 ¥#,##0，中文表头列宽需要更宽的自适应。

## 设计要点
- 样式预设（StylePreset）把"财务建模的行业惯例"固化成可复用对象，
  与后续 Skill 层的 references/style.md 共享同一份规范（单一事实源）。
- 图表用 openpyxl 原生 chart，写入后**保留在工作簿中**，
  但注意：openpyxl 写入的图表在重算后保真度问题需在 Skill 层提醒。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

# ---------------------------------------------------------------------------
# 中文场景配色与字体（与 Skill 层 references/style.md 保持一致的单一事实源）
# ---------------------------------------------------------------------------
# 注：财务建模颜色语义是行业通行惯例（非某家专有），此处为独立表述。
STYLE_SEMANTICS = {
    "input": {"rgb": "0000FF", "meaning": "硬编码输入 / 情景开关"},
    "formula": {"rgb": "000000", "meaning": "公式计算"},
    "cross_sheet": {"rgb": "008000", "meaning": "跨工作表引用"},
    "cross_file": {"rgb": "FF0000", "meaning": "跨文件引用"},
    "assumption_bg": {"rgb": "FFFF00", "meaning": "关键假设 / 需用户填写"},
}

# 中文字体栈：优先雅黑（屏幕阅读），备选宋体（打印/正式）
CJK_FONTS = ["微软雅黑", "Microsoft YaHei", "宋体", "SimSun"]

# 常用中文数字格式
NUMBER_FORMATS = {
    "cny": "¥#,##0",                 # 人民币整数
    "cny2": "¥#,##0.00",             # 人民币两位小数
    "cny_dash": "¥#,##0;(¥#,##0);-",  # 零显示为短横
    "percent": "0.0%",               # 百分比（存小数）
    "multiple": "0.0x",              # 估值倍数
    "int": "#,##0",
    "date": "yyyy-mm-dd",
}


def _ref(ws, cell_range: str) -> Reference:
    """把区域字符串（如 "B2:B5"）构造成 Reference。

    避开 Reference(range_string=...) 要求 "表名!A1:B2" 且中文表名需引号的限制。
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return Reference(
        ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
    )


def _font(name: str | None = None, size: int = 11, bold: bool = False,
          color: str | None = None):
    return Font(name=name or CJK_FONTS[0], size=size, bold=bold,
                color=color or "000000")


def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# 中文样式：一键美化表头 + 列宽自适应 + 冻结
# ---------------------------------------------------------------------------
def apply_chinese_style(
    file_path: Path,
    sheet: str | None = None,
    header_row: int = 1,
    freeze_header: bool = True,
    auto_width: bool = True,
    font_name: str | None = None,
) -> dict:
    """给工作表套用中文场景样式：中文字体表头、自动列宽、冻结首行、细边框。

    参数：
    - header_row：表头所在行（默认第 1 行）
    - auto_width：按内容长度自适应列宽（中文按 2 个字符宽度计算）
    - freeze_header：冻结表头行，便于滚动查看
    """
    wb = load_workbook(filename=str(file_path))
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_name = ws.title
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        # 表头：中文字体 + 加粗 + 浅灰底 + 居中
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        for c in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = _font(font_name, size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()

        # 数据区：中文字体 + 细边框
        for r in range(header_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.font is None or cell.font.name != (font_name or CJK_FONTS[0]):
                    cell.font = _font(font_name, size=11)
                cell.border = _thin_border()

        # 列宽自适应：中文字符按宽度 2 计算
        if auto_width:
            for c in range(1, max_col + 1):
                longest = 0
                for r in range(header_row, min(max_row, header_row + 200) + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    text = str(v)
                    # 中文/全角按 2 宽，ASCII 按 1 宽
                    width = sum(2 if ord(ch) > 127 else 1 for ch in text)
                    longest = max(longest, width)
                ws.column_dimensions[get_column_letter(c)].width = min(
                    max(longest + 2, 8), 50
                )

        if freeze_header:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        wb.save(str(file_path))
        return {
            "file": str(file_path),
            "sheet": sheet_name,
            "styled_rows": max_row,
            "styled_columns": max_col,
            "font": font_name or CJK_FONTS[0],
            "auto_width": auto_width,
            "frozen_header": freeze_header,
        }
    finally:
        wb.close()


def set_number_format(
    file_path: Path,
    sheet: str | None,
    cell_range: str,
    number_format: str,
) -> dict:
    """给区域套数字格式。

    number_format 可直接传内置名（cny / percent / date 等，见 NUMBER_FORMATS）
    或自定义格式码（如 "#,##0.00"）。
    """
    fmt = NUMBER_FORMATS.get(number_format, number_format)
    wb = load_workbook(filename=str(file_path))
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        count = 0
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = fmt
                count += 1
        wb.save(str(file_path))
        return {
            "file": str(file_path),
            "sheet": ws.title,
            "range": cell_range,
            "number_format": fmt,
            "cells_updated": count,
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 图表（官方 skill 与竞品 knorq 都缺失的能力）
# ---------------------------------------------------------------------------
def add_chart(
    file_path: Path,
    sheet: str | None,
    chart_type: str = "bar",
    data_range: str = "",
    categories_range: str | None = None,
    title: str = "",
    anchor_cell: str = "E2",
    width: int = 15,
    height: int = 8,
) -> dict:
    """在工作表中插入图表。

    参数：
    - chart_type：bar / line / pie / scatter
    - data_range：数据区域（含 series 数据，如 "B2:B10"）
    - categories_range：分类轴区域（如 "A2:A10"）
    - anchor_cell：图表左上角锚定单元格
    """
    ct = (chart_type or "bar").lower()
    if ct not in ("bar", "line", "pie", "scatter"):
        raise ValueError(f"不支持的图表类型 '{chart_type}'。可用：bar / line / pie / scatter")

    wb = load_workbook(filename=str(file_path))
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_name = ws.title

        # 坑：Reference 的 range_string 必须是 "sheetname!A1:B2" 形式，
        # 纯区域字符串（如 "B2:B5"）会抛 ValueError，且中文表名还需引号包裹。
        # 改用 range_boundaries 解析成行列参数，既避开表名转义，又支持任意区域。
        data = _ref(ws, data_range)

        chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        if ct == "scatter":
            from openpyxl.chart import ScatterChart

            chart = ScatterChart()
        else:
            chart = chart_classes[ct]()

        chart.title = title or None
        chart.width = width
        chart.height = height

        cats = _ref(ws, categories_range) if categories_range else None

        if ct == "scatter":
            chart.series.append(Series(data, cats, title_from_data=False))
        else:
            chart.add_data(data, titles_from_data=False)
            if cats is not None:
                chart.set_categories(cats)

        ws.add_chart(chart, anchor_cell)
        wb.save(str(file_path))
        return {
            "file": str(file_path),
            "sheet": sheet_name,
            "chart_type": ct,
            "data_range": data_range,
            "categories_range": categories_range,
            "title": title,
            "anchor": anchor_cell,
            "note": (
                "图表已写入。提示：openpyxl 写入的图表在 LibreOffice/Excel 打开后"
                "显示效果可能需微调；若后续执行 recalculate，请重开文件确认图表保真度。"
            ),
        }
    finally:
        wb.close()
