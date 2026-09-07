#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
可交互透视表（真 PivotTable）—— 通过手动构造 OOXML 部件实现。

## 为什么需要这个模块

openpyxl **只能保留已有的** PivotTable，无法创建新的。但 xlsx 本质是
zip + XML，因此可以手动构造透视表所需的 OOXML 部件并注入，从而生成
Excel 可识别的**真·可交互透视表**（带字段拖拽、展开折叠、刷新）。

## 需要注入的部件

```
xl/pivotCache/pivotCacheDefinition1.xml      # 缓存定义（字段 + 共享项）
xl/pivotCache/pivotCacheRecords1.xml         # 缓存记录（每行数据）
xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels
xl/pivotTables/pivotTable1.xml               # 透视表布局（行/列/值/页字段）
```

并同步更新：
- `[Content_Types].xml`      —— 声明新部件的 ContentType
- `xl/workbook.xml`          —— 注册 <pivotCaches>
- `xl/_rels/workbook.xml.rels` —— 指向 pivotCacheDefinition
- `xl/worksheets/_rels/sheetN.xml.rels` —— 把 pivotTable 挂到目标 sheet（**关键**，缺了 Excel 不认）

## 诚实标注（重要限制）

1. **后续用 openpyxl 保存会导致透视表丢失** —— openpyxl 不支持写回 pivot 部件。
   因此：本模块注入透视表后，若再用 duduExcel 其他写工具（write_cells 等）
   修改该文件，透视表会消失。建议**最后一步**再生成透视表。
2. **不支持**：字段分组（日期分组/数字区间）、计算字段/计算项、切片器（Slicer）、
   时间线（Timeline）、多数据源、数据模型（Power Pivot）。
3. LibreOffice 对 PivotTable 的支持有限（能识别但交互弱于 Excel）。
"""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# 聚合函数 → OOXML subtotal 属性
AGG_MAP = {
    "sum": "sum",
    "count": "count",
    "average": "average",
    "mean": "average",
    "avg": "average",
    "min": "min",
    "max": "max",
    "product": "product",
    "count_nums": "countNums",
    "stddev": "stdDev",
    "var": "var",
}

# 中文默认名前缀
AGG_LABEL = {
    "sum": "求和项",
    "count": "计数项",
    "average": "平均值项",
    "mean": "平均值项",
    "avg": "平均值项",
    "min": "最小值项",
    "max": "最大值项",
}


class PivotError(Exception):
    """透视表构造/注入失败。"""


def _esc(text: Any) -> str:
    """XML 属性/文本转义。"""
    s = "" if text is None else str(text)
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _is_number(v: Any) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    return False


# ---------------------------------------------------------------------------
# 1. 读取源数据
# ---------------------------------------------------------------------------
def _read_source(file_path: Path, sheet: str | None):
    """读取源表，返回 (sheet_name, header, rows)。"""
    wb = load_workbook(filename=str(file_path), data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_name = ws.title
        all_rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not all_rows:
        raise PivotError("源工作表为空")
    header = [("" if h is None else str(h)) for h in all_rows[0]]
    data = [list(r) for r in all_rows[1:] if any(c is not None for c in r)]
    return sheet_name, header, data


def _apply_row_filters(header: list[str], data: list[list], filters: list[dict] | None) -> list[list]:
    """按条件过滤行（与 filter_count 同格式）。"""
    if not filters:
        return data
    out = []
    for row in data:
        keep = True
        for f in filters:
            col = f.get("column")
            op = f.get("op", "==")
            val = f.get("value")
            if col not in header:
                continue
            idx = header.index(col)
            cell = row[idx] if idx < len(row) else None
            try:
                if op == "==":
                    keep = cell == val
                elif op == "!=":
                    keep = cell != val
                elif op in (">", "<", ">=", "<="):
                    if cell is None or not _is_number(cell):
                        keep = False
                    else:
                        c, v = float(cell), float(val)
                        keep = {">": c > v, "<": c < v, ">=": c >= v, "<=": c <= v}[op]
                elif op == "in":
                    keep = cell in (val if isinstance(val, list) else [val])
                elif op == "not_in":
                    keep = cell not in (val if isinstance(val, list) else [val])
                elif op == "contains":
                    keep = (val or "") in ("" if cell is None else str(cell))
                elif op == "is_null":
                    keep = cell is None
                elif op == "not_null":
                    keep = cell is not None
            except Exception:
                keep = False
            if not keep:
                break
        if keep:
            out.append(row)
    return out


# ---------------------------------------------------------------------------
# 2. 构造缓存：字段共享项 + 记录
# ---------------------------------------------------------------------------
def _build_cache(header: list[str], data: list[list]):
    """为每列构造 sharedItems 与记录索引。

    约定（与 Excel 生成方式一致）：
    - 文本/其他类型：进 sharedItems，记录中用 <x v="索引"/>
    - 数值类型：记录中直接写 <n v="值"/>，sharedItems 仍列出唯一值以兼容
    """
    n_cols = len(header)
    shared: list[list[str]] = [[] for _ in range(n_cols)]      # 每列的共享项（文本）
    index_of: list[dict] = [{} for _ in range(n_cols)]         # 值 -> 共享项索引
    numeric_col: list[bool] = [True] * n_cols

    for row in data:
        for i in range(n_cols):
            v = row[i] if i < len(row) else None
            if v is None or v == "":
                continue
            if not _is_number(v):
                numeric_col[i] = False

    # 收集文本列的共享项
    for row in data:
        for i in range(n_cols):
            if numeric_col[i]:
                continue
            v = row[i] if i < len(row) else None
            key = "" if v is None else str(v)
            if key not in index_of[i]:
                index_of[i][key] = len(shared[i])
                shared[i].append(key)

    # 数值列也收集唯一值（用于 sharedItems 声明，便于 Excel 显示）
    numeric_unique: list[list[float]] = [[] for _ in range(n_cols)]
    for i in range(n_cols):
        if not numeric_col[i]:
            continue
        seen = set()
        for row in data:
            v = row[i] if i < len(row) else None
            if _is_number(v):
                f = float(v)
                if f not in seen:
                    seen.add(f)
                    numeric_unique[i].append(f)

    # 构造记录
    records: list[str] = []
    for row in data:
        parts = []
        for i in range(n_cols):
            v = row[i] if i < len(row) else None
            if v is None or v == "":
                # 缺失值：文本列用 <m/>（missing），数值列也用 <m/>
                parts.append("<m/>")
            elif numeric_col[i]:
                parts.append(f'<n v="{_esc(v)}"/>')
            else:
                key = str(v)
                idx = index_of[i].get(key)
                parts.append("<m/>" if idx is None else f'<x v="{idx}"/>')
        records.append("<r>" + "".join(parts) + "</r>")

    return shared, numeric_col, numeric_unique, records


def _cache_definition_xml(header, shared, numeric_col, numeric_unique, sheet_name, n_rows, n_cols):
    fields = []
    for i, name in enumerate(header):
        if numeric_col[i]:
            items = "".join(f'<n v="{_esc(v)}"/>' for v in numeric_unique[i])
            contains = ' containsNumber="1"'
            # min/max 便于 Excel 显示
            if numeric_unique[i]:
                mn, mx = min(numeric_unique[i]), max(numeric_unique[i])
                contains += f' minValue="{_esc(mn)}" maxValue="{_esc(mx)}"'
            fields.append(
                f'<cacheField name="{_esc(name)}" numFmtId="0">'
                f'<sharedItems{contains} count="{len(numeric_unique[i])}">{items}</sharedItems>'
                f"</cacheField>"
            )
        else:
            items = "".join(f'<s v="{_esc(v)}"/>' for v in shared[i])
            has_semi = ' containsSemiMixedTypes="0"' if any(
                re.match(r"^\s*$", v) for v in shared[i]) else ""
            fields.append(
                f'<cacheField name="{_esc(name)}" numFmtId="0">'
                f'<sharedItems{has_semi} count="{len(shared[i])}">{items}</sharedItems>'
                f"</cacheField>"
            )
    ref = f"A1:{_col_letter(max(n_cols,1))}{n_rows + 1}"   # 含表头
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'r:id="rIdPivotCacheRec" recordCount="{n_rows}" createdVersion="6" refreshedVersion="6" '
        'minRefreshableVersion="3" refreshOnLoad="1">\n'
        ' <cacheSource type="worksheet">\n'
        f'  <worksheetSource ref="{ref}" sheet="{_esc(sheet_name)}"/>\n'
        ' </cacheSource>\n'
        f' <cacheFields count="{len(header)}">' + "".join(fields) + "</cacheFields>\n"
        "</pivotCacheDefinition>"
    )


def _cache_records_xml(records):
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'count="{len(records)}">' + "".join(records) + "</pivotCacheRecords>"
    )


def _col_letter(idx: int) -> str:
    """1 -> A"""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s or "A"


# ---------------------------------------------------------------------------
# 3. 构造 pivotTableDefinition
# ---------------------------------------------------------------------------
def _pivot_table_xml(header, rows, columns, values, page_fields, agg_func,
                     shared, numeric_col, location_ref, pivot_name):
    row_idx = [header.index(c) for c in rows if c in header]
    col_idx = [header.index(c) for c in (columns or []) if c in header]
    page_idx = [header.index(c) for c in (page_fields or []) if c in header]
    val_idx = [header.index(c) for c in values if c in header]

    if not val_idx:
        raise PivotError(f"值字段不存在于源表：{values}。可用列：{', '.join(header)}")

    subtotal = AGG_MAP.get((agg_func or "sum").lower(), "sum")
    label = AGG_LABEL.get((agg_func or "sum").lower(), "求和项")

    # pivotFields：按列顺序
    pf = []
    for i in range(len(header)):
        attrs = ['showAll="0"']
        if i in row_idx:
            attrs.append('axis="axisRow"')
            attrs.append('outline="1"')
        elif i in col_idx:
            attrs.append('axis="axisCol"')
            attrs.append('outline="1"')
        elif i in page_idx:
            attrs.append('axis="axisPage"')
        if i in val_idx:
            attrs.append('dataField="1"')
        # 行/列/页字段需要 items 列出共享项
        if (i in row_idx or i in col_idx or i in page_idx) and not numeric_col[i] and shared[i]:
            items = '<item t="default"/>' + "".join(
                f'<item x="{k}"/>' for k in range(len(shared[i]))
            )
            pf.append(
                f'<pivotField {" ".join(attrs)}>'
                f'<items count="{len(shared[i]) + 1}">{items}</items>'
                f"</pivotField>"
            )
        else:
            pf.append(f'<pivotField {" ".join(attrs)}/>')

    row_fields = (
        f'<rowFields count="{len(row_idx)}">'
        + "".join(f'<field x="{i}"/>' for i in row_idx)
        + "</rowFields>"
        if row_idx else ""
    )
    col_fields = (
        f'<colFields count="{len(col_idx)}">'
        + "".join(f'<field x="{i}"/>' for i in col_idx)
        + "</colFields>"
        if col_idx else ""
    )
    page_fields_xml = (
        f'<pageFields count="{len(page_idx)}">'
        + "".join(f'<pageField fld="{i}"/>' for i in page_idx)
        + "</pageFields>"
        if page_idx else ""
    )
    data_fields = (
        f'<dataFields count="{len(val_idx)}">'
        + "".join(
            f'<dataField name="{_esc(label)}:{_esc(header[i])}" fld="{i}" '
            f'subtotal="{subtotal}" numFmtId="0"/>'
            for i in val_idx
        )
        + "</dataFields>"
    )

    # 逐段拼接（避免条件表达式混入字符串拼接导致字段丢失）
    segments = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n',
        '<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" ',
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
        f'name="{_esc(pivot_name)}" cacheId="1" applyNumberFormats="0" applyBorderFormats="0" ',
        'applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" ',
        'applyWidthHeightFormats="1" dataOnRows="0" updatedVersion="6" ',
        'minRefreshableVersion="3" showCalcMbrs="1" useAutoFormatting="1" ',
        'itemPrintTitles="1" createdVersion="6" indent="0" outline="1" outlineData="1" ',
        'multipleFieldFilters="0">\n',
        f' <location ref="{location_ref}" firstHeaderRow="1" firstDataRow="1" firstDataCol="0"/>\n',
        f' <pivotFields count="{len(header)}">' + "".join(pf) + "</pivotFields>\n",
    ]
    if row_fields:
        segments.append(f" {row_fields}\n")
    if col_fields:
        segments.append(f" {col_fields}\n")
    if page_fields_xml:
        segments.append(f" {page_fields_xml}\n")
    segments.append(f" {data_fields}\n")
    segments.append("</pivotTableDefinition>")
    return "".join(segments)


# ---------------------------------------------------------------------------
# 4. 注入 zip 并更新各 XML
# ---------------------------------------------------------------------------
def _next_rid(rels_xml: str) -> str:
    """从现有 rels 计算下一个不冲突的 rId。"""
    ids = re.findall(r'Id="rId(\d+)"', rels_xml)
    n = max((int(i) for i in ids), default=0) + 1
    return f"rId{n}"


def _sheet_part_for(workbook_xml: str, workbook_rels: str, sheet_name: str) -> str | None:
    """根据 sheet 名找到其 xml 部件路径（如 worksheets/sheet2.xml）。"""
    m = re.search(r'<sheet[^>]*name="([^"]*)"[^>]*r:id="([^"]*)"', workbook_xml)
    # 遍历所有 sheet 标签，匹配目标名
    for tag in re.findall(r"<sheet\s[^>]*?/>", workbook_xml):
        nm = re.search(r'name="([^"]*)"', tag)
        rid = re.search(r'r:id="([^"]*)"', tag)
        if nm and rid and nm.group(1) == sheet_name:
            target_rid = rid.group(1)
            for rel in re.findall(r"<Relationship\s[^>]*?/>", workbook_rels):
                i = re.search(r'Id="([^"]*)"', rel)
                t = re.search(r'Target="([^"]*)"', rel)
                if i and t and i.group(1) == target_rid:
                    tgt = t.group(1)
                    if tgt.startswith("/xl/"):
                        return tgt[4:]
                    if tgt.startswith("xl/"):
                        return tgt[3:]
                    return tgt
    return None


def create_interactive_pivot(
    file_path: Path,
    source_sheet: str | None = None,
    rows: list[str] | None = None,
    values: list[str] | None = None,
    agg_func: str = "sum",
    columns: list[str] | None = None,
    page_fields: list[str] | None = None,
    filters: list[dict] | None = None,
    target_sheet: str = "透视表",
    location: str = "A3",
    pivot_name: str = "duduPivot",
) -> dict:
    """创建**可交互**透视表（真 PivotTable）。

    参数：
    - rows / columns / values：行 / 列 / 值字段名（值可多列）
    - agg_func：sum / count / average / min / max（默认 sum）
    - page_fields：报表筛选字段（axisPage）
    - filters：行过滤条件（格式同 filter_count）
    - target_sheet：结果写入的工作表（不存在则新建）
    - location：透视表锚定区域左上角（默认 A3）

    返回：构造结果 + 诚实标注的限制说明。
    """
    file_path = Path(file_path)
    rows = rows or []
    values = values or []
    if not values:
        raise PivotError("必须指定至少一个 values（值字段）")
    if (agg_func or "sum").lower() not in AGG_MAP:
        raise PivotError(
            f"不支持的聚合函数 '{agg_func}'。可用：{', '.join(sorted(set(AGG_MAP)))}"
        )

    # 1) 读源数据
    src_name, header, data = _read_source(file_path, source_sheet)
    missing = [c for c in (rows + values + (columns or []) + (page_fields or []))
               if c not in header]
    if missing:
        raise PivotError(f"字段不存在：{missing}。可用列：{', '.join(header)}")

    data = _apply_row_filters(header, data, filters)
    if not data:
        raise PivotError("过滤后没有数据，无法生成透视表")

    # 2) 确保目标 sheet 存在（openpyxl 先建好并保存）
    wb = load_workbook(filename=str(file_path))
    try:
        if target_sheet not in wb.sheetnames:
            wb.create_sheet(target_sheet)
        wb.save(str(file_path))
    finally:
        wb.close()

    # 3) 构造缓存与布局 XML
    shared, numeric_col, numeric_unique, records = _build_cache(header, data)
    cache_def = _cache_definition_xml(
        header, shared, numeric_col, numeric_unique, src_name, len(data), len(header)
    )
    cache_rec = _cache_records_xml(records)
    # 透视表占用区域（给足空间）
    loc_ref = f"{location}:{_col_letter(max(len(header), 2))}100"
    pt_xml = _pivot_table_xml(
        header, rows, columns, values, page_fields, agg_func, shared,
        numeric_col, loc_ref, pivot_name,
    )

    # 4) 重写 zip：注入部件 + 更新各 XML
    tmp = file_path.with_suffix(file_path.suffix + ".pivottmp")
    ct_override = (
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
        '<Override PartName="/xl/pivotTables/pivotTable1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>'
    )

    with zipfile.ZipFile(file_path) as zin:
        workbook_xml = zin.read("xl/workbook.xml").decode("utf-8", "replace")
        wb_rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
        sheet_part = _sheet_part_for(workbook_xml, wb_rels, target_sheet)

        if sheet_part is None:
            raise PivotError(f"无法定位工作表 '{target_sheet}' 对应的 XML 部件")

        sheet_rels_name = (
            f"xl/worksheets/_rels/{sheet_part.split('/')[-1]}.rels"
        )
        # 读取目标 sheet 的 rels（可能不存在）
        try:
            sheet_rels = zin.read(sheet_rels_name).decode("utf-8", "replace")
        except KeyError:
            sheet_rels = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
            )

        cache_rid = _next_rid(wb_rels)
        ptable_rid = _next_rid(sheet_rels)
        # 无论 sheet rels 是否原本存在，都在这里统一注入 pivotTable 关系，
        # 保证 sheet_rels 变量始终是"已注入"的版本（供循环外写出使用）
        sheet_rels = sheet_rels.replace(
            "</Relationships>",
            f'<Relationship Id="{ptable_rid}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
            'Target="../pivotTables/pivotTable1.xml"/></Relationships>',
        )
        # 兜底：若 rels 是自闭合写法（新建模板常见 <Relationships .../>），
        # 上面 replace 匹配不到 </Relationships> 会静默失败 —— 这里先把
        # 自闭合展开为成对标签再注入一次，确保关系一定写进去。
        if "pivotTable" not in sheet_rels:
            sheet_rels = sheet_rels.replace(
                "/>",
                '><Relationship Id="%s" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" '
                'Target="../pivotTables/pivotTable1.xml"/></Relationships>' % ptable_rid,
            )

        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename
                payload = zin.read(name)

                if name == "[Content_Types].xml":
                    s = payload.decode("utf-8")
                    s = s.replace("</Types>", ct_override + "</Types>")
                    payload = s.encode("utf-8")

                elif name == "xl/workbook.xml":
                    s = payload.decode("utf-8")
                    s = s.replace(
                        "</workbook>",
                        f'<pivotCaches><pivotCache cacheId="1" r:id="{cache_rid}"/>'
                        f"</pivotCaches></workbook>",
                    )
                    payload = s.encode("utf-8")

                elif name == "xl/_rels/workbook.xml.rels":
                    s = payload.decode("utf-8")
                    s = s.replace(
                        "</Relationships>",
                        f'<Relationship Id="{cache_rid}" '
                        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" '
                        'Target="pivotCache/pivotCacheDefinition1.xml"/></Relationships>',
                    )
                    payload = s.encode("utf-8")

                if name == sheet_rels_name:
                    # 已存在：直接写"已注入"的版本（覆盖原内容）
                    payload = sheet_rels.encode("utf-8")

                zout.writestr(item, payload)

            # 关键：目标 sheet 的 rels 可能原本不存在（zip 里没有该条目），
            # 上面的循环不会写出它 —— 必须在此显式写入，否则 Excel 找不到
            # 透视表归属，"可交互"就无从谈起。
            existing = set()
            for n in zin.namelist():
                existing.add(n)
            if sheet_rels_name not in existing:
                zout.writestr(sheet_rels_name, sheet_rels)

            # 新增部件
            zout.writestr("xl/pivotCache/pivotCacheDefinition1.xml", cache_def)
            zout.writestr("xl/pivotCache/pivotCacheRecords1.xml", cache_rec)
            zout.writestr(
                "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdPivotCacheRec" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" '
                'Target="pivotCacheRecords1.xml"/>'
                "</Relationships>",
            )
            zout.writestr("xl/pivotTables/pivotTable1.xml", pt_xml)

    # 5) 原子替换
    import os

    os.replace(str(tmp), str(file_path))

    # 6) 校验
    verify = verify_pivot(file_path, target_sheet)

    return {
        "file": str(file_path),
        "source_sheet": src_name,
        "target_sheet": target_sheet,
        "rows": rows,
        "columns": columns or [],
        "values": values,
        "page_fields": page_fields or [],
        "agg_func": agg_func,
        "rows_aggregated": len(data),
        "is_interactive_pivot": verify.get("pivot_count", 0) > 0,
        "verification": verify,
        "note": (
            "已生成**可交互**透视表（真 PivotTable）。"
            if verify.get("pivot_count", 0) > 0
            else "透视表部件已注入，但 openpyxl 未能识别，请检查结构。"
        ),
        "limitations": [
            "⚠️ 后续若用 openpyxl 保存该文件（包括 duduExcel 的 write_cells 等写工具），"
            "透视表会丢失 —— openpyxl 不支持写回 pivot 部件。建议**最后一步**再生成透视表。",
            "不支持：字段分组（日期/数字区间）、计算字段与计算项、切片器、时间线、多数据源、数据模型。",
            "LibreOffice 能识别但交互能力弱于 Excel。",
        ],
    }


def verify_pivot(file_path: Path, sheet: str | None = None) -> dict:
    """校验透视表是否真的生成（openpyxl 视角 + zip 部件视角）。"""
    result: dict[str, Any] = {"pivot_count": 0}

    # openpyxl 视角
    try:
        wb = load_workbook(filename=str(file_path))
        try:
            names = [sheet] if (sheet and sheet in wb.sheetnames) else wb.sheetnames
            total = 0
            detail = {}
            for sn in names:
                n = len(wb[sn]._pivots)
                detail[sn] = n
                total += n
            result["pivot_count"] = total
            result["by_sheet"] = detail
        finally:
            wb.close()
    except Exception as e:
        result["openpyxl_error"] = f"{type(e).__name__}: {e}"

    # zip 部件视角
    try:
        with zipfile.ZipFile(file_path) as z:
            names = z.namelist()
            result["parts"] = {
                "pivotTables": [n for n in names if "pivotTables/" in n],
                "pivotCacheDefinition": [n for n in names if "pivotCacheDefinition" in n],
                "pivotCacheRecords": [n for n in names if "pivotCacheRecords" in n],
                "sheet_rels_with_pivot": [
                    n for n in names
                    if n.startswith("xl/worksheets/_rels/")
                    and "pivotTable" in z.read(n).decode("utf-8", "replace")
                ],
            }
    except Exception as e:
        result["zip_error"] = f"{type(e).__name__}: {e}"

    return result
