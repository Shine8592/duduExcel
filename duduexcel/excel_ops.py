#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
核心 Excel 操作（纯函数，不依赖 MCP，便于单测）。

设计原则（汲取自调研）：
1. 上下文友好：任何"读"都必须能分页/截断，绝不把整表灌进 Agent 上下文
   （学 jwadow/mcp-excel 的 "results, not rows" 与 limit/offset）。
2. 批量优先：写入一次 load/save 完成多单元格（学 knorq 的 bulk 变体，
   它明确要求 "Use these instead of calling the single-target versions in a loop"）。
3. 诚实截断：截断时必须告知真实总数与被隐藏的行数
   （学 Anthropic 官方 recalc.py 的 "locations_truncated" 诚实性）。
4. 规避 openpyxl 已知坑（官方 xlsx skill 总结的六条）：
   - 只读场景用 read_only=True 提升大文件性能
   - 需要公式/值双读时分别以 data_only 开关加载
   - 写入用普通模式（read_only 工作簿不可写）
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# 单次读取默认返回行数上限（防止 Agent 上下文被大表撑爆）
DEFAULT_READ_LIMIT = 200
# 单次写入的单元格数量上限（对齐 knorq 的"范围上限"思路，先设保守值）
MAX_WRITE_CELLS = 100_000

# ---------------------------------------------------------------------------
# 格式语义标记（汲取自 VOYAGER-Inc/excel-vision-mcp）
#
# 它的洞察：作者留下的格式是**语义信号**——删除线常表示"已取消"、
# 黄底表示"待审阅"、高亮表示"重点"。纯文本读取会把它们抹平，
# 于是"已取消"的行和"生效中"的行在模型眼里长得一模一样。
# 关键设计（一并借鉴）：**只在单元格真的用了格式时才附加标记**，
# 朴素表格不产生任何额外 token。
# ---------------------------------------------------------------------------
# 常见底色 RGB → 颜色名（Excel 标准调色板，前两位为 alpha 通道）
_FILL_COLOR_NAMES = {
    "FFFFFF00": "yellow", "FFFF0000": "red", "FF00FF00": "green",
    "FF0000FF": "blue", "FFFFC000": "orange", "FF808080": "gray",
    "FFFFFF99": "light-yellow", "FFC00000": "dark-red",
    "FF92D050": "light-green", "FF00B0F0": "light-blue",
    "FFD9D9D9": "light-gray", "FFFFE699": "light-yellow",
    # 条件格式常用色
    "FFC7CE": "light-red", "FFC6EFCE": "light-green", "FFEB9C": "light-orange",
}
# 字体颜色同理（常见语义色）
_FONT_COLOR_NAMES = {
    "FFFF0000": "red", "FF0000FF": "blue", "FF008000": "green",
    "FFC00000": "dark-red", "FF808080": "gray", "FFFFC000": "orange",
}


def _rgb_str(color) -> str | None:
    """安全取颜色的 RGB 字符串（openpyxl 的 color 可能为 None 或主题色）。"""
    try:
        if color is None:
            return None
        rgb = getattr(color, "rgb", None)
        if isinstance(rgb, str) and rgb:
            return rgb.upper()
        # 主题色/索引色：无法直接映射到 RGB，返回 None 交给调用方忽略
        return None
    except Exception:
        return None


def _format_markers(cell) -> list[str]:
    """提取单元格的格式语义标记。

    返回如 ["B", "S", "HL:yellow", "C:red"]；无特殊格式时返回空列表
    （朴素表格零额外开销）。
    """
    marks: list[str] = []
    try:
        font = cell.font
        if font is not None:
            if font.bold:
                marks.append("B")
            if font.italic:
                marks.append("I")
            if font.strike:            # 删除线：常表示"已取消/作废"
                marks.append("S")
            rgb = _rgb_str(font.color)
            # 排除默认黑色（FF000000 或未设置）
            if rgb and rgb not in ("FF000000", "00000000"):
                name = _FONT_COLOR_NAMES.get(rgb)
                marks.append(f"C:{name}" if name else f"C:{rgb[-6:]}")
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill is not None and getattr(fill, "fill_type", None) == "solid":
            rgb = _rgb_str(fill.fgColor)
            # 排除"无填充"（全 0）与纯白
            if rgb and rgb not in ("00000000", "FFFFFFFF"):
                name = _FILL_COLOR_NAMES.get(rgb)
                marks.append(f"HL:{name}" if name else f"HL:{rgb[-6:]}")
    except Exception:
        pass
    return marks


def _is_hidden_row(ws, row_idx: int) -> bool:
    """判断行是否隐藏（未显式设置维度的行默认不隐藏）。"""
    try:
        dim = ws.row_dimensions.get(row_idx)
        return bool(dim is not None and dim.hidden)
    except Exception:
        return False


def _is_hidden_col(ws, col_idx: int) -> bool:
    """判断列是否隐藏。"""
    try:
        dim = ws.column_dimensions.get(get_column_letter(col_idx))
        return bool(dim is not None and dim.hidden)
    except Exception:
        return False


def hidden_rows_cols(file_path: Path, sheet_name: str) -> tuple[set, set]:
    """直接从 xlsx 的 sheet XML 解析隐藏行/列。

    为什么不用 ws.row_dimensions：openpyxl 在 `read_only=True` 模式下
    **不加载行列维度**（row_dimensions 为空字典），一律返回"未隐藏"，
    导致隐藏内容检测静默失效。这里改用 zipfile + 正则直接读 sheet XML：
    既准确，又不必把整个工作簿加载进内存。
    """
    import re
    import zipfile

    hidden_rows: set[int] = set()
    hidden_cols: set[int] = set()
    try:
        with zipfile.ZipFile(file_path) as z:
            wbxml = z.read("xl/workbook.xml").decode("utf-8", "replace")
            # sheet 名 → r:id（XML 属性顺序不固定，故逐个 tag 提取）
            sheet_map: dict[str, str] = {}
            for tag in re.findall(r"<sheet\s[^>]*?/>", wbxml):
                n = re.search(r'name="([^"]*)"', tag)
                r = re.search(r'r:id="([^"]*)"', tag)
                if n and r:
                    sheet_map[n.group(1)] = r.group(1)

            rid = sheet_map.get(sheet_name)
            if not rid:
                return hidden_rows, hidden_cols

            rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
            target = None
            for tag in re.findall(r"<Relationship\s[^>]*?/>", rels):
                i = re.search(r'Id="([^"]*)"', tag)
                t = re.search(r'Target="([^"]*)"', tag)
                if i and t and i.group(1) == rid:
                    target = t.group(1)
                    break
            if not target:
                return hidden_rows, hidden_cols

            # Target 可能是绝对路径（/xl/...）或相对 xl/ 的路径
            member = target.lstrip("/")
            if not member.startswith("xl/"):
                member = "xl/" + member
            sxml = z.read(member).decode("utf-8", "replace")

            # 行：<row r="3" hidden="1" ...>
            for tag in re.findall(r"<row\s[^>]*?>", sxml):
                if 'hidden="1"' in tag:
                    m = re.search(r'r="(\d+)"', tag)
                    if m:
                        hidden_rows.add(int(m.group(1)))
            # 列：<col min="2" max="2" hidden="1" ...>
            for tag in re.findall(r"<col\s[^>]*?/?>", sxml):
                if 'hidden="1"' in tag:
                    mn = re.search(r'min="(\d+)"', tag)
                    mx = re.search(r'max="(\d+)"', tag)
                    if mn and mx:
                        hidden_cols.update(
                            range(int(mn.group(1)), int(mx.group(1)) + 1)
                        )
    except Exception:
        return hidden_rows, hidden_cols
    return hidden_rows, hidden_cols


def _visible_formula_refs(ws) -> set[str]:
    """收集可见单元格公式里引用的单元格坐标（用于 HIDDEN-REF 例外判断）。

    注意：这是保守的正则提取，可能把函数名中的类坐标片段误计入，
    但误判的代价仅是"多保留一个隐藏单元格"，不会丢失信息。
    """
    import re

    refs: set[str] = set()
    pat = re.compile(r"\$?([A-Z]{1,3})\$?(\d{1,7})")
    try:
        for row in ws.iter_rows():
            for cell in row:
                v = getattr(cell, "value", None)
                if isinstance(v, str) and v.startswith("="):
                    for m in pat.finditer(v):
                        # 过滤无效列号：正则可能把函数名片段误判为坐标，
                        # 需确保列字母能转成合法列索引（>=1）
                        try:
                            if column_index_from_string(m.group(1)) >= 1:
                                refs.add(f"{m.group(1)}{m.group(2)}")
                        except Exception:
                            continue
    except Exception:
        pass
    return refs


def _open(path: Path, data_only: bool = False, read_only: bool = False):
    """统一打开工作簿。xlsm 需保留 VBA，故按后缀决定 keep_vba。"""
    keep_vba = path.suffix.lower() == ".xlsm"
    return load_workbook(
        filename=str(path),
        data_only=data_only,
        read_only=read_only,
        keep_vba=keep_vba,
    )


def _sheet_names(path: Path) -> list[str]:
    wb = _open(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _pick_sheet(wb, sheet: str | None):
    """选择工作表；未指定用活动表；不存在时给出可读错误 + 可用列表。"""
    if sheet is None:
        return wb.active
    if sheet not in wb.sheetnames:
        raise KeyError(
            f"工作表 '{sheet}' 不存在。可用工作表：{', '.join(wb.sheetnames)}"
        )
    return wb[sheet]


# ---------------------------------------------------------------------------
# L1 探查：一次调用替代多次（学 jwadow 的 get_data_profile）
# ---------------------------------------------------------------------------
def inspect_workbook(file_path: Path) -> dict[str, Any]:
    """返回工作簿概览：工作表清单、每张表的行列数、文件大小。

    设计意图：让 Agent 用一次调用搞清楚文件结构，
    避免"先读全表再猜结构"这种把上下文吃光的做法。
    """
    size_bytes = file_path.stat().st_size
    wb = _open(file_path, read_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            # 隐藏行/列：read_only 模式下维度不可用，统一从 XML 解析
            h_rows, h_cols = hidden_rows_cols(file_path, ws.title)
            hidden_rows, hidden_cols = len(h_rows), len(h_cols)
            try:
                merged_count = len(ws.merged_cells.ranges)
            except Exception:
                merged_count = 0
            sheets.append(
                {
                    "name": ws.title,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                    "merged_cells": merged_count,
                    "hidden_rows": hidden_rows,
                    "hidden_columns": hidden_cols,
                }
            )
        return {
            "file": str(file_path),
            "file_size_kb": round(size_bytes / 1024, 1),
            "active_sheet": wb.active.title if wb.active else None,
            "sheet_count": len(sheets),
            "sheets": sheets,
            "embedded_images": count_images(file_path),
        }
    finally:
        wb.close()


def list_images(file_path: Path) -> dict[str, Any]:
    """列出工作簿内嵌图片（零依赖：直接扫描 xlsx 的 xl/media/ 目录）。

    汲取自 excel-vision-mcp —— 它指出其他 Excel MCP "silently drop every
    embedded image"，导致贴在单元格里的流程图永远到不了模型眼前。
    本实现用标准库 zipfile 扫描归档，因此**不需要额外依赖**；
    返回清单（数量/文件名/尺寸）而非图片本体，避免图片 base64 撑爆上下文。
    """
    import zipfile

    if file_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"file": str(file_path), "images": [], "count": 0,
                "note": "仅 .xlsx/.xlsm 支持内嵌图片扫描"}

    items: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(file_path) as z:
            for name in z.namelist():
                if not name.startswith("xl/media/"):
                    continue
                if name.endswith("/"):
                    continue
                info = z.getinfo(name)
                item: dict[str, Any] = {
                    "path": name,
                    "filename": name.split("/")[-1],
                    "size_kb": round(info.file_size / 1024, 1),
                }
                # 尝试读 PNG/JPEG 尺寸（读文件头即可，无需 Pillow）
                dims = _image_size(z.read(name)[:64])
                if dims:
                    item["width"], item["height"] = dims
                items.append(item)
    except zipfile.BadZipFile:
        return {"file": str(file_path), "images": [], "count": 0,
                "note": "文件不是有效的 xlsx 归档"}
    except Exception as e:
        return {"file": str(file_path), "images": [], "count": 0,
                "note": f"扫描失败：{e}"}

    return {
        "file": str(file_path),
        "count": len(items),
        "images": items,
        "note": (
            f"发现 {len(items)} 张内嵌图片（位于 xl/media/）。"
            "本工具只返回清单以避免图片数据撑爆上下文；"
            "如需让模型'看到'图片内容，请把文件解压后由多模态模型读取。"
            if items else "该工作簿没有内嵌图片"
        ),
    }


def _image_size(head: bytes):
    """从文件头解析 PNG/JPEG/GIF/BMP 尺寸（不依赖 Pillow）。"""
    try:
        if head[:8] == b"\x89PNG\r\n\x1a\n" and len(head) >= 24:
            w = int.from_bytes(head[16:20], "big")
            h = int.from_bytes(head[20:24], "big")
            return w, h
        if head[:2] == b"\xff\xd8":  # JPEG：需扫描 SOF 段
            i = 2
            while i < len(head) - 9:
                if head[i] != 0xFF:
                    i += 1
                    continue
                marker = head[i + 1]
                if marker in (0xC0, 0xC1, 0xC2, 0xC3):
                    h = int.from_bytes(head[i + 5 : i + 7], "big")
                    w = int.from_bytes(head[i + 7 : i + 9], "big")
                    return w, h
                i += 2
            return None
        if head[:6] in (b"GIF87a", b"GIF89a") and len(head) >= 10:
            return int.from_bytes(head[6:8], "little"), int.from_bytes(head[8:10], "little")
        if head[:2] == b"BM" and len(head) >= 26:
            return int.from_bytes(head[18:22], "little"), int.from_bytes(head[22:26], "little")
    except Exception:
        pass
    return None


def count_images(file_path: Path) -> int:
    """快速统计内嵌图片数量（供 workbook_info 使用）。"""
    try:
        return list_images(file_path).get("count", 0)
    except Exception:
        return 0


# ---------------------------------------------------------------------------
# L3 读取：分页 + 截断 + 诚实报告
# ---------------------------------------------------------------------------
def read_range(
    file_path: Path,
    sheet: str | None = None,
    cell_range: str | None = None,
    offset: int = 0,
    limit: int = DEFAULT_READ_LIMIT,
    include_formula: bool = True,
    include_format: bool = True,
    include_hidden: bool = False,
) -> dict[str, Any]:
    """读取单元格区域，支持分页、截断、格式语义与隐藏内容处理。

    - offset/limit：分页（limit 默认 200，防止上下文溢出）
    - 返回 total/returned/truncated，截断时明确告知隐藏了多少
    - include_formula=True 时同时给出公式与缓存值（学 knorq 的 read_cell 双字段）
    - include_format=True 时附加格式语义标记 [B]/[I]/[S]/[HL:色]/[C:色]/[M]
      （学 excel-vision-mcp：删除线=已取消、黄底=待审阅，都是作者留下的语义；
        且只在真的用了格式时才附加，朴素表零开销）
    - include_hidden=False（默认）时跳过隐藏行/列——作者藏起来通常意味着
      "不属于要审阅的内容"；但**被可见公式引用**的隐藏单元格仍保留并标记
      [HIDDEN-REF]，因为它的值会影响你看得见的结果（该例外同样学自 excel-vision-mcp）
    """
    if limit <= 0:
        raise ValueError("limit 必须大于 0")
    if offset < 0:
        raise ValueError("offset 不能为负数")

    wb_value = _open(file_path, data_only=True, read_only=True)
    try:
        ws = _pick_sheet(wb_value, sheet)
        sheet_name = ws.title

        # 合并单元格集合（用于 [M] 标记）
        try:
            merged = {str(r) for r in ws.merged_cells.ranges}
        except Exception:
            merged = set()

        # 隐藏行/列：read_only 模式下 row_dimensions 为空，故从 XML 解析
        if include_hidden:
            h_rows: set[int] = set()
            h_cols: set[int] = set()
        else:
            h_rows, h_cols = hidden_rows_cols(file_path, sheet_name)

        # 被可见公式引用的隐藏单元格需要保留（HIDDEN-REF 例外）
        hidden_refs: set[str] = set()
        if not include_hidden and (h_rows or h_cols):
            hidden_refs = _visible_formula_refs(ws)

        if cell_range:
            rows_iter = ws[cell_range]
        else:
            rows_iter = ws.iter_rows()

        all_rows: list[list[Any]] = []
        all_cells: list[list[Any]] = []
        hidden_rows_skipped = 0
        hidden_cols_skipped = 0
        kept_hidden_refs = 0

        for row in rows_iter:
            if not row:
                continue
            r_idx = getattr(row[0], "row", len(all_rows) + 1)

            # 隐藏行处理
            if not include_hidden and r_idx in h_rows:
                # 例外：该行有单元格被可见公式引用 → 保留
                row_refs = {
                    f"{get_column_letter(getattr(c, 'column', 0))}{r_idx}"
                    for c in row
                    if getattr(c, "column", 0) and getattr(c, "column", 0) >= 1
                }
                if not (row_refs & hidden_refs):
                    hidden_rows_skipped += 1
                    continue
                kept_hidden_refs += 1

            vals: list[Any] = []
            cells_row: list[Any] = []
            for c in row:
                c_idx = getattr(c, "column", len(vals) + 1)
                if not include_hidden and c_idx in h_cols:
                    coord = f"{get_column_letter(c_idx)}{r_idx}"
                    if coord not in hidden_refs:
                        hidden_cols_skipped += 1
                        continue
                    kept_hidden_refs += 1
                vals.append(_clean(getattr(c, "value", c)))
                cells_row.append(c)
            all_rows.append(vals)
            all_cells.append(cells_row)

            if len(all_rows) >= offset + limit + 1:
                break
    finally:
        wb_value.close()

    total_rows = len(all_rows)
    page = all_rows[offset : offset + limit]
    page_cells = all_cells[offset : offset + limit]
    has_more = total_rows > offset + len(page)

    # 公式读取（仅在需要时才二次打开文件）
    formulas: dict[str, str] | None = None
    if include_formula:
        formulas = {}
        wb_formula = _open(file_path, data_only=False, read_only=True)
        try:
            ws_f = _pick_sheet(wb_formula, sheet_name)
            src = ws_f[cell_range] if cell_range else ws_f.iter_rows()
            for r_i, row in enumerate(src):
                if r_i < offset:
                    continue
                if r_i >= offset + len(page):
                    break
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell.coordinate] = cell.value
        except Exception:
            formulas = None
        finally:
            wb_formula.close()

    # 组装行文本与结构化数据
    grid: list[list[Any]] = []
    cells_out: dict[str, Any] = {}
    format_map: dict[str, list[str]] = {}
    lines: list[str] = []

    for r_off, (row_vals, row_cells) in enumerate(zip(page, page_cells)):
        real_row = offset + r_off + 1
        parts: list[str] = []
        for c_off, (v, cell) in enumerate(zip(row_vals, row_cells)):
            coord = getattr(cell, "coordinate", None) or \
                    f"{get_column_letter(c_off+1)}{real_row}"
            if v is not None:
                cells_out[coord] = v
            marks: list[str] = []
            if include_format:
                marks = _format_markers(cell)
                if coord in merged:
                    marks.append("M")
                if marks:
                    format_map[coord] = marks
            # 隐藏但被公式引用而保留的单元格
            if not include_hidden and (real_row in h_rows or (c_off + 1) in h_cols):
                marks = marks + ["HIDDEN-REF"]
                format_map[coord] = marks
            if v is not None or marks:
                seg = f"{coord}: {'' if v is None else v}"
                if marks:
                    seg += " " + "".join(f"[{m}]" for m in marks)
                parts.append(seg.strip())
        grid.append(list(row_vals))
        if parts:
            lines.append(" | ".join(parts))

    result: dict[str, Any] = {
        "file": str(file_path),
        "sheet": sheet_name,
        "range": cell_range or "(全表)",
        "offset": offset,
        "returned_rows": len(page),
        "truncated": has_more,
        "grid": grid,
        "cells": cells_out,
    }

    # 格式语义：只有真的有标记时才附上，朴素表零开销
    if format_map:
        result["format_markers"] = format_map
        result["formatted_view"] = lines
        result["marker_legend"] = {
            "B": "粗体", "I": "斜体", "S": "删除线（常表示已取消/作废）",
            "HL:色": "单元格底色/高亮（常表示待审阅、重点）",
            "C:色": "字体颜色", "M": "合并单元格",
            "HIDDEN-REF": "该单元格已隐藏，但因被可见公式引用而保留",
        }
    if formulas:
        result["formulas"] = formulas

    if has_more:
        result["hint"] = (
            f"结果已截断：仅返回 {len(page)} 行（offset={offset}）。"
            f"如需更多请增大 limit 或调整 offset。"
        )

    # 隐藏内容的处理报告（不静默丢弃）
    if not include_hidden and (hidden_rows_skipped or hidden_cols_skipped):
        result["hidden_skipped"] = {
            "rows": hidden_rows_skipped,
            "columns": hidden_cols_skipped,
            "kept_due_to_formula_ref": kept_hidden_refs,
            "note": (
                f"已跳过 {hidden_rows_skipped} 个隐藏行、{hidden_cols_skipped} 个隐藏列"
                f"（作者隐藏通常表示不属于审阅内容）。"
                f"其中 {kept_hidden_refs} 个因被可见公式引用而保留，标记 [HIDDEN-REF]。"
                "如需全部读取请传 include_hidden=true。"
            ),
        }
    return result


def _row_values(row) -> list[Any]:
    """取一行的值列表。

    坑：openpyxl 在任意模式下 iter_rows() / ws[range] 迭代出的都是 Cell 对象
    （read_only 模式为 ReadOnlyCell），必须取 .value 才是单元格内容，
    直接把 Cell 对象当值用会拿到 "<ReadOnlyCell '表名'.A1>" 这种字符串。
    """
    out: list[Any] = []
    for c in row:
        # 单元格对象取 value；若已是原始值（防御性）则直接用
        out.append(_clean(getattr(c, "value", c)))
    return out


def _clean(v: Any) -> Any:
    """清洗不可直接 JSON 序列化的值。"""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        # 官方 skill 提到：公式结果为 "" 时 openpyxl 也会读回 None，这里保持原样
        return v
    return str(v)


# ---------------------------------------------------------------------------
# L3 写入：bulk 批量 + 单次读写周期
# ---------------------------------------------------------------------------
def write_cells(
    file_path: Path,
    sheet: str | None,
    cells: list[dict[str, Any]],
    create_sheet_if_missing: bool = False,
) -> dict[str, Any]:
    """批量写入多个单元格，一次 load/save 完成。

    cells 形如 [{"cell": "A1", "value": "姓名"}, {"cell": "B2", "value": "=SUM(B3:B9)"}]
    - value 以 "=" 开头视为公式（学 knorq 的约定）
    - 支持跨坐标一次性写入，避免 Agent 循环调用单格接口
    """
    if not cells:
        raise ValueError("cells 为空，未提供任何要写入的单元格")
    if len(cells) > MAX_WRITE_CELLS:
        raise ValueError(
            f"单次写入 {len(cells)} 个单元格超过上限 {MAX_WRITE_CELLS}，请分批写入"
        )

    wb = _open(file_path)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        elif create_sheet_if_missing:
            ws = wb.create_sheet(sheet)
        else:
            raise KeyError(
                f"工作表 '{sheet}' 不存在。可用工作表：{', '.join(wb.sheetnames)}"
                "（如需新建请传 create_sheet_if_missing=true）"
            )

        written = []
        for item in cells:
            coord = str(item.get("cell", "")).strip().upper()
            if not coord:
                raise ValueError(f"单元格项缺少 cell 坐标：{item}")
            value = item.get("value")
            # 官方 skill 的坑：空字符串与 None 语义不同，这里按原样写入
            ws[coord] = value
            written.append(
                {"cell": coord, "value": value, "is_formula": _is_formula(value)}
            )

        # 原子保存：写临时文件成功后再替换目标，失败则原文件不受影响
        from duduexcel.safety import atomic_save

        atomic_save(wb, file_path)
        return {
            "file": str(file_path),
            "sheet": ws.title,
            "written_count": len(written),
            "cells": written,
            "note": (
                "注意：公式已写入但尚未重算，读取时可能得到 None；"
                "重算能力将在 M3 阶段提供（recalculate 工具）。"
                if any(c["is_formula"] for c in written)
                else None
            ),
        }
    finally:
        wb.close()


def _is_formula(value: Any) -> bool:
    """判断值是否为 Excel 公式。"""
    return isinstance(value, str) and value.startswith("=")
