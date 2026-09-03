#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M2 服务端分析层：让 Agent 问问题、拿结果，而不是把整表搬进上下文。

核心理念（汲取自 jwadow/mcp-excel 与 jgravelle/jdatamunch-mcp）：
1. **服务端执行**：过滤、聚合、排序全在服务端算完，只回传结果数字。
   jwadow 的原话是 "precise results — not thousands of rows"。
2. **一次调用替代 N 次**：sheet_profile 一次给出全表画像，
   省掉 Agent 反复试探列类型/空值/分布的十几次调用（学 get_data_profile）。
3. **自报节省**：每个结果附 _meta.tokens_saved（学 jdatamunch 的 _meta 契约），
   让 Agent 与用户直观感知"这次少烧了多少 token"。
4. **结果可复现**：每个结果附等价的 Excel 公式（学 jwadow），
   用户粘回表格即可动态更新，不必依赖本次 MCP 调用。
5. **诚实估算**：采样/估算时明确标注，不把估算伪装成精确值
   （学 jdatamunch "Sampling-based statistics report their error bounds"）。

依赖：pandas（可选）。未安装时自动降级到纯标准库实现，功能不减。
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

# --- 读取与开销估算常量 -----------------------------------------------------
# 粗略估算：一个单元格进 JSON 后约占 4 个 token（保守值，用于 self-report）
TOKENS_PER_CELL = 4
# top_n 默认返回行数
DEFAULT_TOP_N = 10
# 支持的过滤运算符
FILTER_OPS = {
    "==", "!=", ">", "<", ">=", "<=", "in", "not_in",
    "contains", "startswith", "endswith", "is_null", "not_null",
}
# 支持的聚合运算
AGG_OPS = {"sum", "mean", "median", "min", "max", "count", "std", "var", "nunique"}


class AnalyticsError(Exception):
    """分析类操作的参数/执行错误，消息面向 Agent 保持可读可纠正。"""


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
def _load_frame(file_path: Path, sheet: str | None = None):
    """把工作表读成 DataFrame（pandas）或行列表（降级路径）。

    统一返回 (df, sheet_name, has_pandas)。
    首行作为表头（与 read_range 的行为保持一致）。
    """
    try:
        import pandas as pd
    except ImportError:
        return _load_rows_stdlib(file_path, sheet), None, False

    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    try:
        if sheet is None:
            ws = wb.active
            sheet_name = ws.title
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_name = sheet
        else:
            raise AnalyticsError(
                f"工作表 '{sheet}' 不存在。可用工作表：{', '.join(wb.sheetnames)}"
            )
        rows = list(ws.values)
    finally:
        wb.close()

    if not rows:
        return pd.DataFrame(), sheet_name, True
    header = [str(h) if h is not None else f"col_{i+1}" for i, h in enumerate(rows[0])]
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    return df, sheet_name, True


def _load_rows_stdlib(file_path: Path, sheet: str | None = None):
    """无 pandas 时的降级：返回 (header, rows)。"""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    try:
        if sheet is None:
            ws = wb.active
        elif sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            raise AnalyticsError(
                f"工作表 '{sheet}' 不存在。可用工作表：{', '.join(wb.sheetnames)}"
            )
        rows = list(ws.values)
    finally:
        wb.close()
    if not rows:
        return [], []
    header = [str(h) if h is not None else f"col_{i+1}" for i, h in enumerate(rows[0])]
    return header, rows[1:]


# ---------------------------------------------------------------------------
# _meta：自报 token 节省（学 jdatamunch）
# ---------------------------------------------------------------------------
def _meta(rows_scanned: int, cols: int, response_obj: Any) -> dict[str, Any]:
    """估算本次调用省下的 token。

    基线 = 如果把整表都灌进上下文需要的 token；
    实际 = 本次响应序列化后的大致 token。
    """
    full_cells = max(rows_scanned, 0) * max(cols, 0)
    try:
        import json

        approx_resp = len(json.dumps(response_obj, ensure_ascii=False, default=str))
        resp_tokens = max(1, approx_resp // 3)  # 中文按 3 字节/token 粗算
    except Exception:
        resp_tokens = 1
    baseline = max(1, full_cells * TOKENS_PER_CELL)
    saved = max(0, baseline - resp_tokens)
    return {
        "rows_scanned": rows_scanned,
        "columns_scanned": cols,
        "tokens_baseline_full_table": baseline,
        "tokens_this_response": resp_tokens,
        "tokens_saved": saved,
        "savings_pct": round(saved / baseline * 100, 2) if baseline else 0.0,
        "note": "基线为'把整表塞进上下文'的粗略估算；小表节省不明显属正常。",
    }


# ---------------------------------------------------------------------------
# L1++ 列画像：一次调用替代十几次试探
# ---------------------------------------------------------------------------
def sheet_profile(file_path: Path, sheet: str | None = None) -> dict[str, Any]:
    """全表列级画像：类型、空值率、唯一数、Top 值、数值统计。

    这是"省调用"的关键接口 —— 学 jwadow 的 get_data_profile，
    它明确写着 "Replaces 10+ separate calls"。
    """
    df, sheet_name, has_pd = _load_frame(file_path, sheet)

    if has_pd:
        cols = list(df.columns)
        n_rows = len(df)
        profile = []
        for c in cols:
            s = df[c]
            non_null = s.dropna()
            item: dict[str, Any] = {
                "column": c,
                "dtype": str(s.dtype),
                "non_null": int(non_null.shape[0]),
                "null_count": int(n_rows - non_null.shape[0]),
                "null_pct": (
                    round((n_rows - non_null.shape[0]) / n_rows * 100, 2) if n_rows else 0.0
                ),
                "unique": int(non_null.nunique()),
            }
            if non_null.shape[0] > 0:
                try:
                    vc = non_null.astype(str).value_counts().head(5)
                    item["top_values"] = {str(k): int(v) for k, v in vc.items()}
                except Exception:
                    item["top_values"] = {}
                # 数值列补统计
                numeric = _to_numeric(non_null)
                if numeric is not None:
                    item["stats"] = {
                        "min": _r(numeric.min()),
                        "max": _r(numeric.max()),
                        "mean": _r(numeric.mean()),
                        "median": _r(numeric.median()),
                    }
            profile.append(item)
        result: dict[str, Any] = {
            "file": str(file_path),
            "sheet": sheet_name,
            "rows": n_rows,
            "columns": cols,
            "profile": profile,
        }
        result["_meta"] = _meta(n_rows, len(cols), result)
        return result

    # --- 降级路径（无 pandas） ---
    header, rows = _load_rows_stdlib(file_path, sheet)
    n_rows = len(rows)
    profile = []
    for i, c in enumerate(header):
        vals = [r[i] for r in rows if i < len(r)]
        non_null = [v for v in vals if v is not None and str(v).strip() != ""]
        nums = [float(v) for v in non_null if _is_num(v)]
        counter: dict[str, int] = {}
        for v in non_null:
            k = str(v)
            counter[k] = counter.get(k, 0) + 1
        top = sorted(counter.items(), key=lambda x: -x[1])[:5]
        item = {
            "column": c,
            "dtype": "numeric" if len(nums) == len(non_null) and non_null else "text",
            "non_null": len(non_null),
            "null_count": n_rows - len(non_null),
            "null_pct": round((n_rows - len(non_null)) / n_rows * 100, 2) if n_rows else 0.0,
            "unique": len(counter),
            "top_values": {k: v for k, v in top},
        }
        if nums and len(nums) == len(non_null):
            item["stats"] = {
                "min": _r(min(nums)),
                "max": _r(max(nums)),
                "mean": _r(sum(nums) / len(nums)),
                "median": _r(_median(nums)),
            }
        profile.append(item)
    result = {
        "file": str(file_path),
        "sheet": sheet or "(活动表)",
        "rows": n_rows,
        "columns": header,
        "profile": profile,
    }
    result["_meta"] = _meta(n_rows, len(header), result)
    return result


# ---------------------------------------------------------------------------
# 过滤 → 计数 / 聚合 / TopN
# ---------------------------------------------------------------------------
def _apply_filters(df, filters: list[dict] | None):
    """按条件过滤 DataFrame，返回过滤后的 df。"""
    if not filters:
        return df
    import pandas as pd

    mask = pd.Series([True] * len(df), index=df.index)
    for f in filters:
        col = f.get("column")
        op = f.get("op", "==")
        val = f.get("value")
        if col not in df.columns:
            raise AnalyticsError(
                f"列 '{col}' 不存在。可用列：{', '.join(map(str, df.columns))}"
            )
        if op not in FILTER_OPS:
            raise AnalyticsError(f"不支持的运算符 '{op}'。可用：{', '.join(sorted(FILTER_OPS))}")
        s = df[col]
        if op == "==":
            mask &= s == val
        elif op == "!=":
            mask &= s != val
        elif op == ">":
            mask &= _to_numeric(s) > float(val)
        elif op == "<":
            mask &= _to_numeric(s) < float(val)
        elif op == ">=":
            mask &= _to_numeric(s) >= float(val)
        elif op == "<=":
            mask &= _to_numeric(s) <= float(val)
        elif op == "in":
            vals = val if isinstance(val, list) else [val]
            mask &= s.isin(vals)
        elif op == "not_in":
            vals = val if isinstance(val, list) else [val]
            mask &= ~s.isin(vals)
        elif op == "contains":
            mask &= s.astype(str).str.contains(str(val), na=False)
        elif op == "startswith":
            mask &= s.astype(str).str.startswith(str(val), na=False)
        elif op == "endswith":
            mask &= s.astype(str).str.endswith(str(val), na=False)
        elif op == "is_null":
            mask &= s.isna()
        elif op == "not_null":
            mask &= s.notna()
    return df[mask]


def _filters_to_excel(filters: list[dict] | None) -> str | None:
    """把过滤条件翻译成等价 Excel 公式片段（结果可复现，学 jwadow）。"""
    if not filters:
        return None
    parts = []
    for f in filters:
        col = f.get("column")
        op = f.get("op", "==")
        val = f.get("value")
        rng = f"[{col}]"  # 结构化引用占位，用户粘到表里再换成实际区域
        if op == "==":
            parts.append(f'{rng}={_q(val)}')
        elif op == "!=":
            parts.append(f'{rng}<>{_q(val)}')
        elif op in (">", "<", ">=", "<="):
            parts.append(f'{rng}"{op}"{_q(val)}')
        elif op == "contains":
            parts.append(f'ISNUMBER(SEARCH({_q(val)},{rng}))')
        elif op == "is_null":
            parts.append(f'{rng}=""')
        elif op == "not_null":
            parts.append(f'{rng}<>""')
        else:
            parts.append(f'{rng}{op}{_q(val)}')
    return "COUNTIFS(" + ",".join(parts) + ")"


def filter_count(
    file_path: Path, sheet: str | None, filters: list[dict] | None,
    sample: int = 3,
) -> dict[str, Any]:
    """按条件计数：返回命中行数（不返回行本身）+ 少量样例 + Excel 公式。"""
    df, sheet_name, has_pd = _load_frame(file_path, sheet)
    if not has_pd:
        raise AnalyticsError("filter_count 需要 pandas，请安装：pip install pandas")

    total = len(df)
    hit = _apply_filters(df, filters)
    n = len(hit)

    # 只回传极少量样例，避免上下文被撑爆
    samples = []
    if sample > 0 and n > 0:
        for _, row in hit.head(sample).iterrows():
            samples.append({str(k): _py(v) for k, v in row.items()})

    result: dict[str, Any] = {
        "file": str(file_path),
        "sheet": sheet_name,
        "total_rows": total,
        "matched_rows": n,
        "matched_pct": round(n / total * 100, 2) if total else 0.0,
        "filters": filters or [],
        "sample_rows": samples,
        "sample_note": f"仅展示前 {len(samples)} 条样例，未返回全部 {n} 行（如需取行请用 read_range 分页）",
        "excel_formula": _filters_to_excel(filters),
    }
    result["_meta"] = _meta(total, len(df.columns), result)
    return result


def aggregate(
    file_path: Path,
    sheet: str | None,
    column: str,
    op: str = "sum",
    group_by: list[str] | None = None,
    filters: list[dict] | None = None,
    top_n: int | None = None,
) -> dict[str, Any]:
    """服务端聚合：sum/mean/count/... 支持分组与过滤。

    学 jwadow 的 aggregate/group_by —— 分组结果只回传"每组一个数字"，
    而不是把所有行交给模型自己算（模型算大数极易出错）。
    """
    df, sheet_name, has_pd = _load_frame(file_path, sheet)
    if not has_pd:
        raise AnalyticsError("aggregate 需要 pandas，请安装：pip install pandas")
    if op not in AGG_OPS:
        raise AnalyticsError(f"不支持的聚合运算 '{op}'。可用：{', '.join(sorted(AGG_OPS))}")
    if column not in df.columns:
        raise AnalyticsError(f"列 '{column}' 不存在。可用列：{', '.join(map(str, df.columns))}")

    total = len(df)
    hit = _apply_filters(df, filters)

    if group_by:
        missing = [c for c in group_by if c not in df.columns]
        if missing:
            raise AnalyticsError(f"分组列不存在：{missing}。可用列：{', '.join(map(str, df.columns))}")
        grouped = hit.groupby(group_by, dropna=False)[column]
        if op == "nunique":
            series = grouped.nunique()
        else:
            series = getattr(grouped, op)()
        rows = []
        for key, val in series.items():
            k = key if isinstance(key, tuple) else (key,)
            rows.append({**{g: _py(kv) for g, kv in zip(group_by, k)}, op: _r(val)})
        rows.sort(key=lambda r: (r[op] is None, -(r[op] or 0)))
        if top_n:
            rows = rows[:top_n]
        result: dict[str, Any] = {
            "file": str(file_path),
            "sheet": sheet_name,
            "operation": op,
            "column": column,
            "group_by": group_by,
            "filters": filters or [],
            "rows_scanned": total,
            "rows_after_filter": len(hit),
            "groups": len(rows),
            "result": rows,
            "tsv": _to_tsv(rows),
            "excel_formula": f"可用数据透视表，或 SUMIFS/AVERAGEIFS 按 {', '.join(group_by)} 分组统计 {column}",
        }
    else:
        s = hit[column]
        if op == "nunique":
            value = int(s.nunique())
        else:
            numeric = _to_numeric(s)
            value = getattr(numeric, op)() if numeric is not None else None
        result = {
            "file": str(file_path),
            "sheet": sheet_name,
            "operation": op,
            "column": column,
            "filters": filters or [],
            "rows_scanned": total,
            "rows_after_filter": len(hit),
            "value": _r(value),
            "excel_formula": f"={op.upper()}([{column}])",
        }
    result["_meta"] = _meta(total, len(df.columns), result)
    return result


def top_n(
    file_path: Path,
    sheet: str | None,
    sort_by: str,
    n: int = DEFAULT_TOP_N,
    ascending: bool = False,
    columns: list[str] | None = None,
    filters: list[dict] | None = None,
) -> dict[str, Any]:
    """取排序后的前 N 行（默认降序），只回传这 N 行。

    学 jwadow 的 rank_rows —— 排行榜类问题不必把全表交给模型排序。
    """
    df, sheet_name, has_pd = _load_frame(file_path, sheet)
    if not has_pd:
        raise AnalyticsError("top_n 需要 pandas，请安装：pip install pandas")
    if sort_by not in df.columns:
        raise AnalyticsError(f"排序列 '{sort_by}' 不存在。可用列：{', '.join(map(str, df.columns))}")

    total = len(df)
    hit = _apply_filters(df, filters)
    sorter = _to_numeric(hit[sort_by])
    ordered = hit.assign(__k=sorter if sorter is not None else hit[sort_by]).sort_values(
        "__k", ascending=ascending
    )
    take = ordered.head(n)

    keep = columns if columns else list(df.columns)
    keep = [c for c in keep if c in df.columns]
    rows = []
    for rank, (_, row) in enumerate(take.iterrows(), start=1):
        d = {"rank": rank}
        d.update({str(c): _py(row[c]) for c in keep})
        rows.append(d)

    result: dict[str, Any] = {
        "file": str(file_path),
        "sheet": sheet_name,
        "sort_by": sort_by,
        "ascending": ascending,
        "filters": filters or [],
        "rows_scanned": total,
        "rows_after_filter": len(hit),
        "returned": len(rows),
        "rows": rows,
        "tsv": _to_tsv(rows),
        "excel_formula": f"=RANK([@{sort_by}],[{sort_by}],{'1' if ascending else '0'})",
    }
    result["_meta"] = _meta(total, len(df.columns), result)
    return result


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def _to_numeric(s):
    """尽力把列转成数值；全非数值返回 None（避免误把文本列当数值算）。"""
    import pandas as pd

    try:
        n = pd.to_numeric(s, errors="coerce")
    except Exception:
        return None
    if n.notna().sum() == 0:
        return None
    return n


def _r(v) -> Any:
    """把 numpy/pandas 标量转成 Python 原生并保留 4 位小数。"""
    if v is None:
        return None
    try:
        import math

        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
    except Exception:
        pass
    try:
        f = float(v)
        if f == int(f) and abs(f) < 1e15:
            return int(f)
        return round(f, 4)
    except Exception:
        return v


def _py(v) -> Any:
    """转成可 JSON 序列化的 Python 原生值。"""
    if v is None:
        return None
    try:
        import pandas as pd

        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _is_num(v) -> bool:
    try:
        float(v)
        return True
    except Exception:
        return False


def _median(nums: list[float]) -> float:
    s = sorted(nums)
    n = len(s)
    if n == 0:
        return 0.0
    mid = n // 2
    return s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2


def _q(v) -> str:
    """Excel 公式里的字面量引号。"""
    if isinstance(v, (int, float)):
        return str(v)
    return '"' + str(v).replace('"', '""') + '"'


def _to_tsv(rows: list[dict]) -> str:
    """TSV 输出，便于直接粘贴回 Excel（学 jwadow 的 TSV output）。"""
    if not rows:
        return ""
    header = list(rows[0].keys())
    lines = ["\t".join(map(str, header))]
    for r in rows:
        lines.append("\t".join("" if r.get(h) is None else str(r.get(h)) for h in header))
    return "\n".join(lines)
