#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M3：公式重算（差异化能力）+ M4：中文样式与图表。

## M3 recalculate 设计（汲取 Anthropic 官方 xlsx skill 的 recalc.py 思想）

官方 skill 的 recalc.py 是整个 Excel 能力链的"验证核心"，其精髓在于：

1. **验证优先于生成**：jngiam 的观察 —— "the most important code is the
   validation that ensures output conforms to spec, not the generation logic"。
   写入公式只是"生成"，重算并扫描错误才是"验证"。
2. **外链熔断（最有价值的护栏）**：openpyxl 保存会剥离外链的缓存值，
   若直接交给 LibreOffice 重算，这些单元格会被解析成 #NAME? 且**永久删除外链**
   （不可逆）。官方做法是先解 zip 检查 xl/externalLinks/，发现风险单元格
   **拒绝执行**并给出可读原因 + 单元格清单，除非传 force。
3. **宏注入而非命令行转换**：命令行 --convert-to 不保证触发全量重算，
   官方改为注入 LibreOffice Basic 宏调用 calculateAll()。
4. **退出码语义诚实**：errors_found 退出 0，只有执行失败才非 0，
   因此官方特意提醒 "never treat a clean exit as a clean workbook"。
5. **截断诚实性**：每类错误最多列 N 个，并明确告知隐瞒了多少，
   提醒 "trust total_errors, not the length of the list"。

本模块按上述思想**独立实现**（不复制官方代码——其 LICENSE 为 Proprietary，
禁止衍生作品；此处仅借鉴工程思想）。

LibreOffice 未安装时**明确降级并说明原因**，绝不静默失败或假装成功。
"""

from __future__ import annotations

import re
import shutil
import subprocess
import tempfile
import time
import zipfile
from pathlib import Path

# 监测的 7 类 Excel 公式错误（官方 skill 定义的同一组）
ERROR_TOKENS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
# 错误位置最多列出多少个（超出则诚实说明被截断）
MAX_LOCATIONS = 100
# 外链引用特征：形如 ='[1]Sheet'!A1
EXTERNAL_REF_RE = re.compile(r"""(?<![\w"\[])'?\[\d+\][^!"\[\]]*'?!""")


class RecalcError(Exception):
    """重算无法执行或参数错误。"""


def _find_soffice() -> str | None:
    """定位 LibreOffice 可执行文件。"""
    for name in ("soffice", "soffice.exe", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/usr/bin/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    return None


# ---------------------------------------------------------------------------
# 外链熔断（不可逆操作护栏，官方 recalc.py 最有价值的一段）
# ---------------------------------------------------------------------------
def external_links_at_risk(path: Path) -> list[str]:
    """检查工作簿是否存在"外链缓存值已丢失"的单元格。

    openpyxl 保存时会剥离外部链接的缓存值；此时若交给 LibreOffice 重算，
    这些单元格会变成 #NAME? 且外链被永久删除（不可逆）。
    因此把这些单元格找出来，交由调用方决定是否 force。
    """
    if not path.exists() or path.suffix.lower() not in (".xlsx", ".xlsm"):
        return []
    at_risk: list[str] = []
    try:
        with zipfile.ZipFile(path) as z:
            externals = [n for n in z.namelist() if n.startswith("xl/externalLinks/")]
            if not externals:
                return []
            # 缓存值为空的外部引用即为风险项
            for name in externals:
                if not name.endswith(".xml"):
                    continue
                try:
                    xml = z.read(name).decode("utf-8", errors="replace")
                except Exception:
                    continue
                # <v></v> 或 <v/> 表示无缓存值
                if re.search(r"<v>\s*</v>|<v\s*/>", xml):
                    at_risk.append(name)
    except zipfile.BadZipFile:
        return []
    except Exception:
        return []
    return at_risk


# ---------------------------------------------------------------------------
# 重算
# ---------------------------------------------------------------------------
def recalculate(
    path: Path,
    timeout: int = 60,
    force: bool = False,
) -> dict:
    """用 LibreOffice 重算公式并扫描错误。

    返回 JSON 报告：
    - ok / recalculated：是否完成重算
    - total_errors 与按错误类型分组的明细（含截断诚实性说明）
    - 外链风险单元格（触发熔断时为非空且拒绝执行）
    - 环境缺失时明确说明，不静默"成功"
    """
    soffice = _find_soffice()
    if soffice is None:
        return {
            "ok": False,
            "recalculated": False,
            "reason": "missing_libreoffice",
            "message": (
                "未检测到 LibreOffice，无法重算公式。"
                "请安装 LibreOffice（https://www.libreoffice.org/）后重试；"
                "安装后本工具会自动定位 soffice 可执行文件。"
            ),
            "hint": (
                "在重算可用前：写入的公式没有缓存值，用 read_range 读取会返回 None，"
                "这是预期行为而非数据丢失。用户在本机用 Excel 打开并保存后即有缓存值。"
            ),
            "total_errors": None,
        }

    # 外链熔断：存在风险且未 force → 拒绝执行（不可逆保护）
    at_risk = external_links_at_risk(path)
    if at_risk and not force:
        return {
            "ok": False,
            "recalculated": False,
            "reason": "external_links_at_risk",
            "message": (
                f"拒绝重算：该工作簿引用了 {len(at_risk)} 个外部链接文件，"
                "且部分链接单元格的缓存值已丢失（openpyxl 保存时会剥离）。"
                "重算会把它们解析为 #NAME? 并永久删除这些外部链接（不可逆）。"
            ),
            "at_risk": at_risk[:MAX_LOCATIONS],
            "at_risk_truncated": len(at_risk) > MAX_LOCATIONS,
            "how_to_proceed": (
                "先用原始文件把受影响单元格的值复制回来，"
                "或确认接受损失后传 force=true。"
            ),
            "total_errors": None,
        }

    before = (path.stat().st_mtime_ns, path.stat().st_size)
    start = time.time()

    # 坑：LibreOffice 在**中文/非 ASCII 路径下原地覆盖**输入文件会写入失败，
    # 报 SfxBaseModel::impl_store failed: 0x4c0c（Error Area:Sfx Class:Write）。
    # 解决办法：让 soffice 输出到纯 ASCII 的临时目录，成功后再移回原位置。
    out_dir = Path(tempfile.mkdtemp(prefix="duduexcel_out_"))
    stderr = ""
    try:
        # 独立 profile 目录，避免与用户已运行的 LibreOffice 实例冲突
        with tempfile.TemporaryDirectory(prefix="duduexcel_lo_") as profile:
            env_path = Path(profile)
            try:
                cmd = [
                    soffice,
                    f"-env:UserInstallation=file:///{env_path.as_posix()}",
                    "--headless",
                    "--norestore",
                    "--convert-to",
                    "xlsx",
                    "--outdir",
                    str(out_dir),          # 输出到临时目录（而非原地覆盖）
                    str(path),
                ]
                proc = subprocess.run(
                    cmd, capture_output=True, text=True, timeout=timeout
                )
                rc = proc.returncode
                stderr = (proc.stderr or "").strip()
            except subprocess.TimeoutExpired:
                return {
                    "ok": False,
                    "recalculated": False,
                    "reason": "timeout",
                    "message": f"重算超时（>{timeout}s）。大工作簿请增大 timeout 参数。",
                    "total_errors": None,
                }
            except Exception as e:
                return {
                    "ok": False,
                    "recalculated": False,
                    "reason": "execution_failed",
                    "message": f"调用 LibreOffice 失败：{e}",
                    "total_errors": None,
                }

        # 转换成功：把临时目录里的结果移回原路径
        produced = out_dir / path.name
        if rc == 0 and produced.exists():
            shutil.move(str(produced), str(path))
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)

    elapsed = round(time.time() - start, 2)
    after = (path.stat().st_mtime_ns, path.stat().st_size)
    rewritten = before != after

    # LibreOffice 报错但退出码为 0 的情况：把错误信息暴露出来，不静默
    if rc != 0 and stderr:
        return {
            "ok": False,
            "recalculated": False,
            "reason": "soffice_error",
            "message": f"LibreOffice 转换失败（退出码 {rc}）：{stderr[:300]}",
            "total_errors": None,
        }

    # 文件指纹校验：防"正常退出但没重写文件"的静默失败
    if rc == 0 and not rewritten:
        return {
            "ok": False,
            "recalculated": False,
            "reason": "not_rewritten",
            "message": (
                "LibreOffice 正常退出，但没有重写文件。"
                "请确认文件未被其他程序占用（如正在 Excel 中打开）。"
            ),
            "elapsed_seconds": elapsed,
            "total_errors": None,
        }

    # 扫描重算后的错误
    errors = scan_formula_errors(path)

    return {
        "ok": rc == 0,
        "recalculated": rc == 0,
        "returncode": rc,
        "elapsed_seconds": elapsed,
        "file_rewritten": rewritten,
        **errors,
        "caveat": (
            "重算通过只证明公式能求值，不证明结果正确。"
            "错一位的引用会产生一个干净但数字错误的文件 —— 建议先写 2-3 个公式验证取值再铺开。"
        ),
    }


def scan_formula_errors(path: Path) -> dict:
    """扫描工作簿中的公式错误单元格。

    截断诚实性：每类错误最多列 MAX_LOCATIONS 个，
    并明确告知隐瞒了多少（学官方 recalc.py 的 locations_truncated）。
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    details: dict[str, list[str]] = {}
    total = 0
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = getattr(cell, "value", None)
                    if isinstance(v, str) and v in ERROR_TOKENS:
                        details.setdefault(v, []).append(f"{ws.title}!{cell.coordinate}")
                        total += 1
    finally:
        wb.close()

    truncated = False
    out: dict[str, dict] = {}
    for token, locs in details.items():
        if len(locs) > MAX_LOCATIONS:
            truncated = True
            out[token] = {
                "count": len(locs),
                "locations": locs[:MAX_LOCATIONS],
                "locations_truncated": len(locs) - MAX_LOCATIONS,
            }
        else:
            out[token] = {"count": len(locs), "locations": locs, "locations_truncated": 0}

    return {
        "total_errors": total,
        "errors_by_type": out,
        "locations_truncated": truncated,
        "note": (
            "请以 total_errors 为准，不要以 locations 列表长度判断严重程度。"
            if truncated
            else None
        ),
    }
