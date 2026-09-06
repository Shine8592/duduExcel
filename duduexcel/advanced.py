#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
M6：透视表、条件格式、多表关联。

补齐此前 README 中列为"待做"的能力，让 duduExcel 覆盖日常表格处理的主要场景。

设计原则（延续前序阶段的调研结论）：
1. **服务端算完只回传结果**：多表关联/比较只返回差异摘要，不把两张表都搬进上下文。
2. **诚实标注实现边界**：openpyxl 原生不支持创建真正的 PivotTable 对象
   （只能保留已有的），因此这里用"分组聚合 + 写回新表"的方式实现等价效果，
   并在返回中明确说明它生成的是**静态汇总表**而非可交互透视表——
   不把等价物伪装成真透视表（学 knorq 诚实列 Known Limitations 的态度）。
3. 条件格式用 openpyxl 原生规则（真实生效，非视觉近似）。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from duduexcel.analytics import _load_frame, _py, _r

# 条件格式类型
COND_TYPES = {
    "data_bar": "数据条（长度表示大小）",
    "color_scale": "色阶（颜色深浅表示大小）",
    "greater_than": "大于阈值时高亮",
    "less_than": "小于阈值时高亮",
    "equal": "等于指定值时高亮",
    "between": "介于两值之间时高亮",
    "contains_text": "包含指定文本时高亮",
    "duplicate": "重复值高亮",
}


class AdvancedError(Exception):
    """高级操作参数或执行错误。"""


# ---------------------------------------------------------------------------
# 透视表（静态汇总表，等价物）
# ---------------------------------------------------------------------------
def create_pivot(
    file_path: Path,
    source_sheet: str | None,
    rows: list[str],
    values: list[str],
    agg_func: str = "sum",
    columns: list[str] | None = None,
    filters: list[dict] | None = None,
    target_sheet: str = "透视表",
) -> dict:
    """生成透视汇总表（写入新工作表）。

    重要说明：openpyxl 无法创建真正可交互的 PivotTable 对象，
    本工具用"分组聚合 + 写回"实现等价的**静态汇总表**。
    若需要可交互透视表，请在 Excel 中基于结果表插入。

    参数：
    - rows：行分组字段（可多列）
    - values：要聚合的数值列
    - agg_func：sum/mean/count/min/max
    - columns：列分组字段（可选，做交叉表）
    - target_sheet：结果写入的工作表名（默认"透视表"）
    """
    df, sheet_name, has_pd = _load_frame(file_path, source_sheet)
    if not has_pd:
        raise AdvancedError("create_pivot 需要 pandas，请安装：pip install pandas")

    missing = [c for c in (rows + values + (columns or [])) if c not in df.columns]
    if missing:
        raise AdvancedError(
            f"列不存在：{missing}。可用列：{', '.join(map(str, df.columns))}"
        )
    if agg_func not in ("sum", "mean", "count", "min", "max"):
        raise AdvancedError(f"不支持的聚合函数 '{agg_func}'。可用：sum/mean/count/min/max")

    # 过滤
    if filters:
        from duduexcel.analytics import _apply_filters

        df = _apply_filters(df, filters)

    # 交叉表：columns 做列维度
    if columns:
        pivot = df.pivot_table(
            index=rows, columns=columns, values=values, aggfunc=agg_func, fill_value=0
        )
    else:
        pivot = df.groupby(rows, dropna=False)[values].agg(agg_func)

    wb = load_workbook(filename=str(file_path))
    try:
        if target_sheet in wb.sheetnames:
            del wb[target_sheet]
        ws = wb.create_sheet(target_sheet)

        # 写入表头
        if columns:
            # 多级列：先写列分组，再写值列名
            ws.append(list(rows) + [f"{v}" for _, v in pivot.columns])
            for idx, row_vals in enumerate(pivot.index, start=2):
                key = row_vals if isinstance(row_vals, tuple) else (row_vals,)
                ws.append(list(key) + [_r(v) for v in pivot.iloc[idx - 2].tolist()])
        else:
            ws.append(list(rows) + list(values))
            for idx, row_vals in enumerate(pivot.index, start=2):
                key = row_vals if isinstance(row_vals, tuple) else (row_vals,)
                ws.append(list(key) + [_r(v) for _, v in pivot.iloc[idx - 2].items()])

        # 表头加粗
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        wb.save(str(file_path))

        return {
            "file": str(file_path),
            "source_sheet": sheet_name,
            "target_sheet": target_sheet,
            "rows": rows,
            "columns": columns,
            "values": values,
            "agg_func": agg_func,
            "result_rows": int(len(pivot)),
            "result_columns": int(pivot.shape[1] if hasattr(pivot, "shape") else len(values)),
            "is_interactive_pivot": False,
            "note": (
                "生成的是**静态汇总表**（openpyxl 无法创建真正可交互的 PivotTable）。"
                "需要可交互透视表时，请在 Excel 中基于本表插入。"
            ),
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 条件格式（openpyxl 原生规则，真实生效）
# ---------------------------------------------------------------------------
def add_conditional_format(
    file_path: Path,
    sheet: str | None,
    cell_range: str,
    cond_type: str,
    value: float | str | None = None,
    value2: float | str | None = None,
    color: str = "FFC7CE",
) -> dict:
    """给区域添加条件格式。

    cond_type 可选（见 COND_TYPES）：
    - data_bar / color_scale：无需 value
    - greater_than / less_than / equal：需要 value
    - between：需要 value 与 value2
    - contains_text：需要 value（文本）
    - duplicate：无需 value（高亮重复值）
    """
    ct = (cond_type or "").lower()
    if ct not in COND_TYPES:
        raise AdvancedError(
            f"不支持的条件格式类型 '{cond_type}'。可用：{', '.join(COND_TYPES)}"
        )

    wb = load_workbook(filename=str(file_path))
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_name = ws.title

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        if ct == "data_bar":
            rule = DataBarRule(start_type="min", end_type="max", color="638EC6")
        elif ct == "color_scale":
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        elif ct == "greater_than":
            if value is None:
                raise AdvancedError("greater_than 需要 value 参数")
            rule = CellIsRule(operator="greaterThan", formula=[str(value)], fill=fill)
        elif ct == "less_than":
            if value is None:
                raise AdvancedError("less_than 需要 value 参数")
            rule = CellIsRule(operator="lessThan", formula=[str(value)], fill=fill)
        elif ct == "equal":
            if value is None:
                raise AdvancedError("equal 需要 value 参数")
            rule = CellIsRule(operator="equal", formula=[str(value)], fill=fill)
        elif ct == "between":
            if value is None or value2 is None:
                raise AdvancedError("between 需要 value 与 value2 两个参数")
            rule = CellIsRule(operator="between", formula=[str(value), str(value2)], fill=fill)
        elif ct == "contains_text":
            if value is None:
                raise AdvancedError("contains_text 需要 value 参数")
            from openpyxl.formatting.rule import FormulaRule

            first = cell_range.split(":")[0]
            col = "".join(ch for ch in first if ch.isalpha())
            row = "".join(ch for ch in first if ch.isdigit())
            rule = FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{value}",{col}{row}))'], fill=fill
            )
        else:  # duplicate
            from openpyxl.formatting.rule import FormulaRule

            first = cell_range.split(":")[0]
            col = "".join(ch for ch in first if ch.isalpha())
            row = "".join(ch for ch in first if ch.isdigit())
            last = cell_range.split(":")[-1] if ":" in cell_range else first
            last_row = "".join(ch for ch in last if ch.isdigit())
            rule = FormulaRule(
                formula=[
                    f"COUNTIF(${col}${row}:${col}${last_row},{col}{row})>1"
                ],
                fill=fill,
            )

        ws.conditional_formatting.add(cell_range, rule)
        wb.save(str(file_path))
        return {
            "file": str(file_path),
            "sheet": sheet_name,
            "range": cell_range,
            "cond_type": ct,
            "description": COND_TYPES[ct],
            "value": value,
            "value2": value2,
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 条件格式读取
# ---------------------------------------------------------------------------
def list_conditional_formats(file_path: Path, sheet: str | None = None) -> dict:
    """读取工作表中已有的条件格式规则。

    说明：调研时竞品 knorq 的 Known Limitations 写着"不支持条件格式"，
    官方文档也常称 openpyxl 读取条件格式受限；但**实测是可以完整读回的**
    （范围 / 类型 / 运算符 / 阈值 / 填充色 / 优先级都能拿到），因此这里补上读取能力，
    形成"写入 + 读取"的闭环。

    用途：
    - 接手一张别人的表时，先看清它埋了哪些规则（哪些格子会自动变红/变色）
    - 修改前确认不破坏既有规则
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path))
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        sheet_name = ws.title

        rules: list[dict[str, Any]] = []
        try:
            for rng in ws.conditional_formatting:
                for rule in rng.rules:
                    item: dict[str, Any] = {
                        "range": str(rng.sqref),
                        "type": rule.type,
                        "operator": getattr(rule, "operator", None),
                        "formula": list(getattr(rule, "formula", None) or []),
                        "priority": getattr(rule, "priority", None),
                    }
                    # 填充色（高亮类规则才有 dxf）
                    dxf = getattr(rule, "dxf", None)
                    if dxf is not None:
                        fill = getattr(dxf, "fill", None)
                        if fill is not None:
                            color = getattr(getattr(fill, "bgColor", None), "rgb", None)
                            if color:
                                item["fill_color"] = color

                    # 关键：构造完必须加入结果列表（此前漏掉这行导致永远返回 0 条）
                    rules.append(item)

            # 按优先级排序，便于阅读
            rules.sort(key=lambda r: (r.get("priority") is None, r.get("priority") or 0))
        except Exception as e:
            return {
                "file": str(file_path),
                "sheet": sheet_name,
                "count": 0,
                "rules": [],
                "note": f"读取条件格式失败：{e}",
            }

        return {
            "file": str(file_path),
            "sheet": sheet_name,
            "count": len(rules),
            "rules": rules,
            "note": (
                f"该工作表共有 {len(rules)} 条条件格式规则。修改这些区域前请先确认不会破坏既有规则。"
                if rules else "该工作表没有条件格式规则"
            ),
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 多表关联：比较 / 连接
# ---------------------------------------------------------------------------
def compare_sheets(
    file_path: Path,
    sheet1: str,
    sheet2: str,
    key_column: str,
    compare_columns: list[str] | None = None,
    max_diff: int = 50,
) -> dict:
    """按关键列比对两个工作表的差异（只回传差异摘要，不回传整表）。

    用途：版本对比、变更检测、对账（学 jwadow 的 compare_sheets）。

    返回：仅在表1/仅在表2/值有差异 三类统计 + 最多 max_diff 条差异明细。
    """
    df1, s1, has_pd = _load_frame(file_path, sheet1)
    df2, s2, _ = _load_frame(file_path, sheet2)
    if not has_pd:
        raise AdvancedError("compare_sheets 需要 pandas，请安装：pip install pandas")
    if key_column not in df1.columns or key_column not in df2.columns:
        raise AdvancedError(
            f"关键列 '{key_column}' 必须在两个表中都存在。"
            f"{s1} 的列：{', '.join(map(str, df1.columns))}；"
            f"{s2} 的列：{', '.join(map(str, df2.columns))}"
        )

    compare_columns = compare_columns or [
        c for c in df1.columns if c in df2.columns and c != key_column
    ]

    m1 = df1.set_index(key_column)
    m2 = df2.set_index(key_column)
    keys1, keys2 = set(m1.index), set(m2.index)

    only1 = sorted(keys1 - keys2, key=str)
    only2 = sorted(keys2 - keys1, key=str)
    common = sorted(keys1 & keys2, key=str)

    diffs = []
    for k in common:
        for col in compare_columns:
            v1 = _py(m1.at[k, col]) if col in m1.columns else None
            v2 = _py(m2.at[k, col]) if col in m2.columns else None
            if v1 != v2:
                diffs.append({"key": _py(k), "column": col, "left": v1, "right": v2})

    truncated = len(diffs) > max_diff
    return {
        "file": str(file_path),
        "left_sheet": s1,
        "right_sheet": s2,
        "key_column": key_column,
        "compared_columns": compare_columns,
        "left_rows": len(df1),
        "right_rows": len(df2),
        "only_in_left": len(only1),
        "only_in_right": len(only2),
        "value_differences": len(diffs),
        "only_in_left_sample": [_py(k) for k in only1[:max_diff]],
        "only_in_right_sample": [_py(k) for k in only2[:max_diff]],
        "differences": diffs[:max_diff],
        "differences_truncated": truncated,
        "note": (
            f"差异明细仅展示前 {max_diff} 条（共 {len(diffs)} 条）。"
            "请以 value_differences 为准判断差异规模。"
            if truncated
            else None
        ),
    }


def join_sheets(
    file_path: Path,
    left_sheet: str,
    right_sheet: str,
    on: str,
    how: str = "left",
    columns: list[str] | None = None,
    limit: int = 20,
) -> dict:
    """关联两个工作表（类似 SQL JOIN），只回传前 limit 行结果。

    参数：
    - on：关联键列名（两表都要有）
    - how：left/right/inner/outer（默认 left）
    - columns：只返回这些列（省 token），省略返回全部
    - limit：返回行数上限（默认 20）
    """
    df1, s1, has_pd = _load_frame(file_path, left_sheet)
    df2, s2, _ = _load_frame(file_path, right_sheet)
    if not has_pd:
        raise AdvancedError("join_sheets 需要 pandas，请安装：pip install pandas")
    if on not in df1.columns or on not in df2.columns:
        raise AdvancedError(
            f"关联键 '{on}' 必须在两个表中都存在。"
            f"{s1} 的列：{', '.join(map(str, df1.columns))}；"
            f"{s2} 的列：{', '.join(map(str, df2.columns))}"
        )
    if how not in ("left", "right", "inner", "outer"):
        raise AdvancedError(f"不支持的关联方式 '{how}'。可用：left/right/inner/outer")

    merged = df1.merge(df2, on=on, how=how, suffixes=("_left", "_right"))
    total = len(merged)

    keep = [c for c in (columns or []) if c in merged.columns]
    view = merged[keep] if keep else merged
    rows = [
        {str(k): _py(v) for k, v in rec.items()}
        for _, rec in view.head(limit).iterrows()
    ]

    from duduexcel.analytics import _meta, _to_tsv

    result = {
        "file": str(file_path),
        "left_sheet": s1,
        "right_sheet": s2,
        "on": on,
        "how": how,
        "total_rows": total,
        "returned": len(rows),
        "truncated": total > len(rows),
        "columns": list(view.columns),
        "rows": rows,
        "tsv": _to_tsv(rows),
        "note": (
            f"仅返回前 {len(rows)} 行（共 {total} 行）。如需更多请用 read_range 读取结果表。"
            if total > len(rows)
            else None
        ),
    }
    result["_meta"] = _meta(total, len(view.columns), result)
    return result
