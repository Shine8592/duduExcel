#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""检查 LibreOffice 是否安装完成；若已就绪则跑通重算闭环验证。"""
import subprocess
import sys
import time
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent.parent))

from duduexcel import recalc

XL = Path(__file__).parent.parent / "tests" / "_tmp_recalc.xlsx"


def check_only() -> bool:
    soffice = recalc._find_soffice()
    if soffice:
        print(f"✅ LibreOffice 已安装: {soffice}")
        return True
    print("⏳ LibreOffice 尚未就绪（安装可能仍在进行）")
    return False


def verify_recalc() -> bool:
    """端到端验证重算闭环：写公式 → 重算 → 读到缓存值。"""
    from openpyxl import Workbook, load_workbook

    if XL.exists():
        XL.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "计算"
    ws.append(["单价", "数量", "金额"])
    ws.append([100, 3, "=A2*B2"])
    ws.append([50, 4, "=A3*B3"])
    ws.append([None, None, "=SUM(C2:C3)"])
    wb.save(XL)

    print("\n=== 重算前（openpyxl 写的公式无缓存值）===")
    wb0 = load_workbook(XL, data_only=True)
    print("  C2 =", wb0["计算"]["C2"].value, "（预期 None）")
    wb0.close()

    print("\n=== 执行 recalculate ===")
    r = recalc.recalculate(XL, timeout=120)
    print("  ok:", r.get("ok"))
    print("  recalculated:", r.get("recalculated"))
    print("  耗时:", r.get("elapsed_seconds"), "秒")
    print("  文件被重写:", r.get("file_rewritten"))
    print("  total_errors:", r.get("total_errors"))
    if r.get("reason"):
        print("  reason:", r.get("reason"))
        print("  message:", r.get("message"))

    if not r.get("recalculated"):
        print("\n❌ 重算未成功，跳过结果校验")
        return False

    print("\n=== 重算后（应读到计算结果）===")
    wb1 = load_workbook(XL, data_only=True)
    c2, c3, c4 = (wb1["计算"][c].value for c in ("C2", "C3", "C4"))
    print("  C2 =", c2, "（预期 300）")
    print("  C3 =", c3, "（预期 200）")
    print("  C4 =", c4, "（预期 500）")
    wb1.close()

    ok = c2 == 300 and c3 == 200 and c4 == 500
    print("\n" + ("✅ 重算闭环验证通过！公式已产生真实缓存值" if ok else "❌ 结果与预期不符"))

    if XL.exists():
        XL.unlink()
    bak = Path(str(XL) + ".bak")
    if bak.exists():
        bak.unlink()
    return ok


if __name__ == "__main__":
    ready = check_only()
    if ready:
        sys.exit(0 if verify_recalc() else 1)
    sys.exit(2)